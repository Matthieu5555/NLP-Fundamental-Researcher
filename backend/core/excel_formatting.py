"""
Professional formatting layer for the Excel financial model.

Single entry point: apply_professional_formatting(workbook, registry, model_audit)
Called at the end of generate_excel() to apply conditional formatting, data
validation, tab colors, an Error Checks sheet with live formula-based integrity
checks, and cell comments from the model auditor.
"""

import logging
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Tab colors (hex without '#')
# ---------------------------------------------------------------------------
TAB_COLORS = {
    "Cover": "808080",
    "Instructions": "808080",
    "Summary": "1F4E79",
    "Conviction": "1F4E79",
    "Assumptions": "4472C4",
    "Financials": "FFFFFF",
    "DCF Model": "ED7D31",
    "Sensitivity": "ED7D31",
    "Scenarios": "7030A0",
    "Earnings Model": "A5A5A5",
    "Comparable Companies": "A5A5A5",
    "Precedent Transactions": "A5A5A5",
    "Checks": "FF0000",
}

# Fills for conditional formatting
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

THIN_BORDER = Border(bottom=Side(style="thin", color="D0D0D0"))
CHECK_FONT = Font(name="Calibri", size=11)
CHECK_FONT_BOLD = Font(name="Calibri", bold=True, size=11)


# ---------------------------------------------------------------------------
# Tab colors
# ---------------------------------------------------------------------------

def _apply_tab_colors(wb: Workbook) -> None:
    for ws in wb.worksheets:
        color = TAB_COLORS.get(ws.title)
        if color:
            ws.sheet_properties.tabColor = color


# ---------------------------------------------------------------------------
# Data validation on Assumptions sheet
# ---------------------------------------------------------------------------

def _apply_data_validation(wb: Workbook) -> None:
    """Add input constraints to Assumptions cells so analysts stay in sane ranges."""
    ws = wb["Assumptions"] if "Assumptions" in wb.sheetnames else None
    if not ws:
        return

    label_validations = {
        "Revenue Growth": DataValidation(
            type="decimal", operator="between",
            formula1="-0.5", formula2="1.0",
            allow_blank=True,
        ),
        "Gross Margin": DataValidation(
            type="decimal", operator="between",
            formula1="0", formula2="1.0",
            allow_blank=True,
        ),
        "OpEx % Revenue": DataValidation(
            type="decimal", operator="between",
            formula1="0", formula2="1.0",
            allow_blank=True,
        ),
        "CapEx % Revenue": DataValidation(
            type="decimal", operator="between",
            formula1="0", formula2="1.0",
            allow_blank=True,
        ),
        "Risk-Free Rate": DataValidation(
            type="decimal", operator="between",
            formula1="0", formula2="0.10",
            allow_blank=True,
        ),
        "Beta": DataValidation(
            type="decimal", operator="between",
            formula1="0", formula2="3.0",
            allow_blank=True,
        ),
        "Equity Risk Premium": DataValidation(
            type="decimal", operator="between",
            formula1="0", formula2="0.15",
            allow_blank=True,
        ),
        "Weight Equity": DataValidation(
            type="decimal", operator="between",
            formula1="0.10", formula2="1.0",
            allow_blank=True,
        ),
        "Tax Rate": DataValidation(
            type="decimal", operator="between",
            formula1="0", formula2="0.50",
            allow_blank=True,
        ),
        "Terminal Growth Rate": DataValidation(
            type="decimal", operator="between",
            formula1="-0.02", formula2="0.08",
            allow_blank=True,
        ),
        "D&A % Revenue": DataValidation(
            type="decimal", operator="between",
            formula1="0", formula2="0.30",
            allow_blank=True,
        ),
        "SBC % Revenue": DataValidation(
            type="decimal", operator="between",
            formula1="0", formula2="0.30",
            allow_blank=True,
        ),
        "NWC % Rev Delta": DataValidation(
            type="decimal", operator="between",
            formula1="-0.30", formula2="0.30",
            allow_blank=True,
        ),
    }

    for label, dv in label_validations.items():
        dv.promptTitle = label
        dv.prompt = f"Editable assumption: {label}"

    for row_idx in range(1, ws.max_row + 1):
        label_cell = ws.cell(row=row_idx, column=1).value
        if not label_cell:
            continue
        label_str = str(label_cell).strip()

        for match_label, dv in label_validations.items():
            if label_str == match_label:
                for col_idx in range(2, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        dv.add(cell)
                ws.add_data_validation(dv)
                break


# ---------------------------------------------------------------------------
# Conditional formatting
# ---------------------------------------------------------------------------

def _apply_conditional_formatting(wb: Workbook) -> None:
    """Apply conditional formatting rules to key cells."""
    if "Sensitivity" in wb.sheetnames:
        ws = wb["Sensitivity"]
        for row_idx in range(1, ws.max_row + 1):
            label = ws.cell(row=row_idx, column=1).value
            if label and "WACC" in str(label) and "TGR" in str(label):
                _apply_grid_color_scale(ws, row_idx + 2, 2, ws.max_row, ws.max_column)
                break


def _apply_grid_color_scale(ws, start_row, start_col, end_row, end_col) -> None:
    """Apply a 3-color scale (red-white-green) to a rectangular data range."""
    range_str = (
        f"{get_column_letter(start_col)}{start_row}:"
        f"{get_column_letter(end_col)}{end_row}"
    )
    rule = ColorScaleRule(
        start_type="min", start_color="FFC7CE",
        mid_type="percentile", mid_value=50, mid_color="FFFFFF",
        end_type="max", end_color="C6EFCE",
    )
    ws.conditional_formatting.add(range_str, rule)


# ---------------------------------------------------------------------------
# Error Checks sheet with live formula-based integrity checks
# ---------------------------------------------------------------------------

_SEVERITY_FILLS = {
    "critical": RED_FILL,
    "warning": YELLOW_FILL,
    "info": BLUE_FILL,
}


def _build_checks_sheet(wb: Workbook, model_audit: Dict[str, Any]) -> None:
    """Build an Error Checks sheet with live formula-based integrity checks.

    Each structural check outputs OK (green) or ERROR (red) via formulas that
    reference named ranges and sheet cells. Model auditor warnings are appended
    below as informational rows.
    """
    ws = wb.create_sheet("Checks")

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=14, color="C00000")
    section_font = Font(name="Calibri", bold=True, size=12, color="C00000")

    ws.cell(row=1, column=1, value="Error Checks Dashboard").font = title_font

    # --- Master status cell (updated at the end with a formula) ---
    master_row = 2
    ws.cell(row=master_row, column=1, value="Overall Status").font = CHECK_FONT_BOLD

    # --- Live structural checks ---
    row = 4
    ws.cell(row=row, column=1, value="Structural Integrity Checks").font = section_font
    row += 1

    for col, val in enumerate(["Check", "Status", "Detail"], 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = header_font
        cell.fill = header_fill
    header_row = row
    row += 1

    check_status_cells: List[str] = []

    # Check 1: TGR < WACC (Gordon Growth validity)
    ws.cell(row=row, column=1, value="TGR < WACC (Gordon Growth)").font = CHECK_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    if "Assumptions" in wb.sheetnames:
        formula = '=IF(IFERROR(TGR>=WACC,FALSE),"ERROR: TGR >= WACC","OK")'
    else:
        formula = "OK"
    ws.cell(row=row, column=2, value=formula).font = CHECK_FONT
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=3, value="Terminal Growth Rate must be below WACC for Gordon Growth to converge").font = CHECK_FONT
    ws.cell(row=row, column=3).border = THIN_BORDER
    check_status_cells.append(f"B{row}")
    row += 1

    # Check 2: Scenario probabilities sum to ~100%
    if "Scenarios" in wb.sheetnames:
        sc_ws = wb["Scenarios"]
        prob_cells = []
        for r in range(1, sc_ws.max_row + 1):
            label = sc_ws.cell(row=r, column=1).value
            if label and str(label).strip().lower() in ("bull", "base", "bear"):
                prob_cells.append(f"'Scenarios'!B{r}")

        ws.cell(row=row, column=1, value="Scenario probabilities sum to 100%").font = CHECK_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        if len(prob_cells) >= 2:
            sum_expr = "+".join(prob_cells)
            formula = f'=IF(ABS({sum_expr}-1)<0.01,"OK","ERROR: Probabilities sum to "&TEXT({sum_expr},"0%"))'
            ws.cell(row=row, column=2, value=formula).font = CHECK_FONT
        else:
            ws.cell(row=row, column=2, value="N/A").font = CHECK_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value="Bull + Base + Bear probabilities should total 100%").font = CHECK_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        check_status_cells.append(f"B{row}")
        row += 1

    # Check 3: WACC in reasonable range (3%-20%)
    if "Assumptions" in wb.sheetnames:
        ws.cell(row=row, column=1, value="WACC in reasonable range (3%-20%)").font = CHECK_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        formula = '=IF(IFERROR(AND(WACC>=0.03,WACC<=0.20),FALSE),"OK","WARNING: WACC="&TEXT(WACC,"0.0%"))'
        ws.cell(row=row, column=2, value=formula).font = CHECK_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value="WACC outside 3-20% range is unusual for most equities").font = CHECK_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        check_status_cells.append(f"B{row}")
        row += 1

    # Check 4: Terminal value is not > 90% of enterprise value
    if "DCF Model" in wb.sheetnames:
        ws.cell(row=row, column=1, value="Terminal value < 90% of EV").font = CHECK_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        formula = (
            '=IF(IFERROR(PVTerminalValue/EnterpriseValue>0.9,FALSE),'
            '"WARNING: TV is "&TEXT(PVTerminalValue/EnterpriseValue,"0%")&" of EV","OK")'
        )
        ws.cell(row=row, column=2, value=formula).font = CHECK_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value="Terminal value exceeding 90% of EV suggests projection period is too short").font = CHECK_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        check_status_cells.append(f"B{row}")
        row += 1

    # Check 5: Fair value is positive
    if "DCF Model" in wb.sheetnames:
        ws.cell(row=row, column=1, value="Fair value is positive").font = CHECK_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        formula = '=IF(IFERROR(FairValue>0,FALSE),"OK","ERROR: Negative fair value")'
        ws.cell(row=row, column=2, value=formula).font = CHECK_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value="Negative implied equity value indicates a structural issue in the model").font = CHECK_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        check_status_cells.append(f"B{row}")
        row += 1

    # Check 6: Enterprise value is positive
    if "DCF Model" in wb.sheetnames:
        ws.cell(row=row, column=1, value="Enterprise value is positive").font = CHECK_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        formula = '=IF(IFERROR(EnterpriseValue>0,FALSE),"OK","ERROR: Negative EV")'
        ws.cell(row=row, column=2, value=formula).font = CHECK_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value="Negative EV suggests negative FCFs exceed terminal value").font = CHECK_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        check_status_cells.append(f"B{row}")
        row += 1

    # Check 7: Tax rate in reasonable range (10%-40%)
    if "Assumptions" in wb.sheetnames:
        has_tax = any(
            dn.name == "TaxRate"
            for dn in wb.defined_names.values()
        )
        if has_tax:
            ws.cell(row=row, column=1, value="Tax rate in range (10%-40%)").font = CHECK_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            formula = '=IF(IFERROR(AND(TaxRate>=0.10,TaxRate<=0.40),FALSE),"OK","WARNING: TaxRate="&TEXT(TaxRate,"0.0%"))'
            ws.cell(row=row, column=2, value=formula).font = CHECK_FONT
            ws.cell(row=row, column=2).border = THIN_BORDER
            ws.cell(row=row, column=3, value="Tax rate outside 10-40% is unusual for most jurisdictions").font = CHECK_FONT
            ws.cell(row=row, column=3).border = THIN_BORDER
            check_status_cells.append(f"B{row}")
            row += 1

    # Check 8: Growth rates decelerating
    if "Assumptions" in wb.sheetnames:
        growth_names = sorted(
            [dn.name for dn in wb.defined_names.values() if dn.name.startswith("GrowthY")],
        )
        if len(growth_names) >= 2:
            ws.cell(row=row, column=1, value="Growth rates decelerating").font = CHECK_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            pairs = [f"{growth_names[i]}>={growth_names[i+1]}" for i in range(len(growth_names) - 1)]
            formula = '=IF(AND(' + ','.join(pairs) + '),"OK","WARNING: Growth not decelerating")'
            ws.cell(row=row, column=2, value=formula).font = CHECK_FONT
            ws.cell(row=row, column=2).border = THIN_BORDER
            ws.cell(row=row, column=3, value="Revenue growth typically decelerates as the company matures").font = CHECK_FONT
            ws.cell(row=row, column=3).border = THIN_BORDER
            check_status_cells.append(f"B{row}")
            row += 1

    # Check 9: TGR below GDP proxy (5%)
    if "Assumptions" in wb.sheetnames:
        ws.cell(row=row, column=1, value="TGR below GDP proxy (5%)").font = CHECK_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        formula = '=IF(IFERROR(TGR<0.05,FALSE),"OK","WARNING: TGR exceeds 5% GDP proxy")'
        ws.cell(row=row, column=2, value=formula).font = CHECK_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value="Terminal growth above long-run GDP growth is generally unsustainable").font = CHECK_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        check_status_cells.append(f"B{row}")
        row += 1

    # Conditional formatting: OK = green, ERROR/WARNING = red
    status_range = f"B{header_row + 1}:B{row - 1}"
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(formula=[f'LEFT(B{header_row + 1},2)="OK"'], fill=GREEN_FILL),
    )
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(formula=[f'LEFT(B{header_row + 1},5)="ERROR"'], fill=RED_FILL),
    )
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(formula=[f'LEFT(B{header_row + 1},7)="WARNING"'], fill=YELLOW_FILL),
    )

    # Master status formula: count ERRORs across all check cells
    if check_status_cells:
        countif_parts = ",".join(
            f'COUNTIF({c},"ERROR*")' for c in check_status_cells
        )
        master_formula = f'=IF({countif_parts}>0,"ERRORS DETECTED","ALL CHECKS PASS")'
        # Simplify: sum all COUNTIF results
        countif_sum = "+".join(f'COUNTIF({c},"ERROR*")' for c in check_status_cells)
        master_formula = f'=IF(({countif_sum})>0,"ERRORS DETECTED","ALL CHECKS PASS")'
        ws.cell(row=master_row, column=2, value=master_formula).font = Font(
            name="Calibri", bold=True, size=14,
        )
        ws.conditional_formatting.add(
            f"B{master_row}",
            FormulaRule(formula=[f'B{master_row}="ALL CHECKS PASS"'], fill=GREEN_FILL),
        )
        ws.conditional_formatting.add(
            f"B{master_row}",
            FormulaRule(formula=[f'B{master_row}="ERRORS DETECTED"'], fill=RED_FILL),
        )

    # --- Model auditor warnings (informational) ---
    row += 1
    warnings = model_audit.get("warnings", []) if model_audit else []
    if warnings:
        ws.cell(row=row, column=1, value="Model Auditor Findings").font = section_font
        row += 1
        for col, val in enumerate(["Severity", "Section", "Finding"], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        for w in warnings:
            sev = w.get("severity", "info")
            fill = _SEVERITY_FILLS.get(sev, BLUE_FILL)
            ws.cell(row=row, column=1, value=sev).font = CHECK_FONT
            ws.cell(row=row, column=2, value=w.get("section", "")).font = CHECK_FONT
            ws.cell(row=row, column=3, value=w.get("message", "")).font = CHECK_FONT
            for c in range(1, 4):
                ws.cell(row=row, column=c).fill = fill
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

    # Auto-width
    for col_idx in range(1, 4):
        max_len = 10
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)


# ---------------------------------------------------------------------------
# Cell comments from assumption_notes
# ---------------------------------------------------------------------------

def _apply_assumption_comments(wb: Workbook, model_audit: Dict[str, Any]) -> None:
    """Attach model auditor assumption notes as cell comments on Assumptions sheet."""
    if "Assumptions" not in wb.sheetnames:
        return
    notes = model_audit.get("assumption_notes", {}) if model_audit else {}
    if not notes:
        return

    ws = wb["Assumptions"]

    key_to_label = {
        "wacc": "WACC",
        "terminal_growth": "Terminal Growth Rate",
        "base_revenue": "Base Revenue",
        "beta": "Beta",
        "risk_free_rate": "Risk-Free Rate",
        "equity_risk_premium": "Equity Risk Premium",
    }

    label_to_row: Dict[str, int] = {}
    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val:
            label_to_row[str(val).strip()] = row_idx

    for note_key, note_text in notes.items():
        label = key_to_label.get(note_key)
        if not label or label not in label_to_row:
            continue
        target_row = label_to_row[label]
        cell = ws.cell(row=target_row, column=2)
        cell.comment = Comment(note_text, "Model Auditor")


# ---------------------------------------------------------------------------
# Historical/projected delineation
# ---------------------------------------------------------------------------

def _apply_estimate_delineation(wb: Workbook) -> None:
    """Add a left border on the first estimate column to visually separate
    historical actuals from projected estimates.

    Detects estimate columns by looking for 'E' suffix in year headers (FY2025E).
    """
    estimate_border = Border(left=Side(style="medium", color="000000"))

    for sheet_name in ("Financials", "Earnings Model"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Scan header rows for the first column ending in 'E'
        first_est_col = None
        for row_idx in range(1, min(ws.max_row + 1, 10)):
            for col_idx in range(2, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and str(val).strip().endswith("E"):
                    first_est_col = col_idx
                    break
            if first_est_col:
                break

        if not first_est_col:
            continue

        # Apply left border to the entire estimate transition column
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=first_est_col)
            existing = cell.border
            cell.border = Border(
                left=Side(style="medium", color="000000"),
                right=existing.right,
                top=existing.top,
                bottom=existing.bottom,
            )


# ---------------------------------------------------------------------------
# Row grouping on DCF and Financials
# ---------------------------------------------------------------------------

def _apply_row_grouping(wb: Workbook) -> None:
    """Add collapsible row groups on detail sections so users can expand/collapse.

    Groups detail rows under subtotals and totals. Looks for subtotal/total
    labels and groups the rows above them.
    """
    # DCF: group Revenue through SBC Add-back (detail) under Free Cash Flow (total)
    if "DCF Model" in wb.sheetnames:
        ws = wb["DCF Model"]
        _group_rows_between_labels(ws, "Revenue", "Free Cash Flow")
        _group_rows_between_labels(ws, "Final Year FCF", "PV of Terminal Value")
        _group_rows_between_labels(ws, "PV of FCFs", "Enterprise Value")

    # Financials: group line items under section headers
    if "Financials" in wb.sheetnames:
        ws = wb["Financials"]
        _group_rows_between_labels(ws, "Revenue", "Net Margin")
        _group_rows_between_labels(ws, "Total Assets", "Current Ratio")
        _group_rows_between_labels(ws, "Operating Cash Flow", "FCF Margin")


def _group_rows_between_labels(ws, start_label: str, end_label: str) -> None:
    """Group rows between two label strings (exclusive of both endpoints)."""
    start_row = None
    end_row = None
    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if not val:
            continue
        val_str = str(val).strip()
        if val_str == start_label and start_row is None:
            start_row = row_idx
        elif val_str == end_label and start_row is not None:
            end_row = row_idx
            break

    if start_row and end_row and end_row > start_row + 1:
        ws.row_dimensions.group(start_row + 1, end_row - 1, outline_level=1)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def _apply_formula_comments(wb: Workbook, dcf: Optional[Dict[str, Any]]) -> None:
    """Attach explanatory comments to key DCF formula cells.

    Scans the DCF Model sheet for label matches and attaches Comment objects
    explaining the underlying financial formula. Also attaches wacc_rationale
    as a comment on the WACC cell on Assumptions when available.
    """
    if "DCF Model" not in wb.sheetnames:
        return

    formula_annotations = {
        "Terminal Value": "Gordon Growth Model: TV = FCF_final * (1+g) / (WACC - g)",
        "Discount Factor": "1/(1+WACC)^n, discounting each year's cash flow to present",
        "Enterprise Value": "PV(FCFs) + PV(Terminal Value)",
        "Equity Value": "EV less net debt and minority interest",
        "Fair Value / Share": "Equity Value / Shares Outstanding",
    }

    ws = wb["DCF Model"]
    for row_idx in range(1, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value
        if not label:
            continue
        label_str = str(label).strip()
        for match_label, explanation in formula_annotations.items():
            if label_str == match_label:
                # Annotate the first formula cell in that row (column 2 or 3)
                for col in (2, 3):
                    cell = ws.cell(row=row_idx, column=col)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        cell.comment = Comment(explanation, "Formula Guide")
                        break
                break

    # Attach wacc_rationale on Assumptions WACC cell if not already commented
    if dcf and "Assumptions" in wb.sheetnames:
        wacc_rationale = (dcf.get("assumptions") or {}).get("wacc_rationale", "")
        if wacc_rationale:
            assumptions_ws = wb["Assumptions"]
            for row_idx in range(1, assumptions_ws.max_row + 1):
                label = assumptions_ws.cell(row=row_idx, column=1).value
                if label and str(label).strip() == "WACC":
                    cell = assumptions_ws.cell(row=row_idx, column=2)
                    if not cell.comment:
                        cell.comment = Comment(wacc_rationale, "Model Auditor")
                    break


def _apply_section_commentary(wb: Workbook, model_audit: Optional[Dict[str, Any]]) -> None:
    """Attach model_audit.section_commentary as comments on each sheet's title cell.

    Maps section keys to sheet names and writes the commentary string as a
    Comment on cell A1 of each matching sheet.
    """
    if not model_audit:
        return
    commentary = model_audit.get("section_commentary", {})
    if not commentary:
        return

    section_to_sheet = {
        "dcf": "DCF Model",
        "scenarios": "Scenarios",
        "comps": "Comparable Companies",
        "precedents": "Precedent Transactions",
        "earnings_model": "Earnings Model",
    }

    for section_key, sheet_name in section_to_sheet.items():
        text = commentary.get(section_key)
        if not text or sheet_name not in wb.sheetnames:
            continue
        cell = wb[sheet_name].cell(row=1, column=1)
        if not cell.comment:
            cell.comment = Comment(text, "Model Auditor")


def apply_professional_formatting(
    wb: Workbook,
    model_audit: Optional[Dict[str, Any]] = None,
    dcf: Optional[Dict[str, Any]] = None,
) -> None:
    """Apply all professional formatting to a completed workbook.

    Called at the end of generate_excel() after all sheets are built.
    """
    _apply_tab_colors(wb)
    _apply_data_validation(wb)
    _apply_conditional_formatting(wb)
    _apply_estimate_delineation(wb)
    _apply_row_grouping(wb)
    _build_checks_sheet(wb, model_audit or {})
    _apply_assumption_comments(wb, model_audit or {})
    _apply_formula_comments(wb, dcf)
    _apply_section_commentary(wb, model_audit)
