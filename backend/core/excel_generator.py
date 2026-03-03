"""
Excel Financial Model Export.

Generates a CFA-quality XLSX workbook with formula-driven sheets.
The Assumptions sheet holds all editable inputs (blue cells); downstream
sheets reference those cells via cross-sheet formulas so an analyst can
change WACC or growth rates and watch the DCF recalculate.

Single entry point: generate_excel() -> bytes

Sensitivity grids are pre-computed (embedding 25 full DCF runs in Excel
formulas is impractical for the decomposed model). All other valuation
math uses live formulas.
"""

import io
import logging
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Protection, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

logger = logging.getLogger(__name__)

# =============================================================================
# STYLE CONSTANTS
# =============================================================================
# Impact of changing: alters all header rows and branded elements in the workbook.
# Colors pulled from the shared palette so Excel stays in sync when theme changes.
from backend.core.colors import PALETTE as _PALETTE


def _hex_strip(color: str) -> str:
    """Strip '#' prefix from hex color for openpyxl (expects 'RRGGBB' not '#RRGGBB')."""
    return color.lstrip("#")


HEADER_COLOR = _hex_strip(_PALETTE.brand.gradient_from)
ACCENT_COLOR = _hex_strip(_PALETTE.semantic.warning)
SUCCESS_COLOR = _hex_strip(_PALETTE.semantic.success_light)
DANGER_COLOR = _hex_strip(_PALETTE.semantic.danger_light)

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
ACCENT_FILL = PatternFill(start_color=ACCENT_COLOR, end_color=ACCENT_COLOR, fill_type="solid")
ACCENT_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

LABEL_FONT = Font(name="Calibri", bold=True, size=11)
VALUE_FONT = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color=HEADER_COLOR)
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color=HEADER_COLOR)

# IB color conventions for formula auditing
INPUT_FONT = Font(name="Calibri", size=11, color="0000CC")        # Blue: editable assumption
FORMULA_FONT = Font(name="Calibri", size=11, color="000000")      # Black: same-sheet formula
LINK_FONT = Font(name="Calibri", size=11, color="008000")         # Green: cross-sheet reference
LABEL_GRAY_FONT = Font(name="Calibri", size=11, color="808080")   # Gray: row labels
RATIONALE_FONT = Font(name="Calibri", size=9, italic=True, color="808080")  # Gray italic: rationale text
ITALIC_LABEL_FONT = Font(name="Calibri", italic=True, size=11)             # Italic: derived ratios

THIN_BORDER = Border(bottom=Side(style="thin", color="D0D0D0"))
SUBTOTAL_BORDER = Border(bottom=Side(style="thin", color="000000"))
TOTAL_BORDER = Border(
    top=Side(style="thin", color="000000"),
    bottom=Side(style="double", color="000000"),
)

UNLOCKED = Protection(locked=False)

# Number formats — parenthetical negatives per IB convention.
# Three commas divide by 1B (each comma = /1000). Two commas = 1M.
# Positive;Negative pattern: negatives render in (brackets), never minus signs.
FMT_USD = '$#,##0.00_);($#,##0.00)'
FMT_USD_INNER = '#,##0.00_);(#,##0.00)'       # No $ sign, for interior rows
FMT_USD_WHOLE = '$#,##0_);($#,##0)'
FMT_PCT = '0.0%'
FMT_PCT_DISPLAY = '0.0"%"'
FMT_RATIO = '0.00"x"'
FMT_NUMBER = '#,##0_);(#,##0)'
FMT_BILLIONS = '#,##0.0,,,"B"_);(#,##0.0,,,"B")'
FMT_MILLIONS = '#,##0.0,,"M"_);(#,##0.0,,"M")'

# Section border used instead of blank rows between sections (preserves Ctrl+Arrow)
SECTION_BORDER = Border(bottom=Side(style="medium", color="000000"))


# =============================================================================
# NAMED RANGE REGISTRY
# =============================================================================

class NamedRangeRegistry:
    """Centralizes named range creation so formula-writing code uses logical
    names instead of hardcoded cell addresses.

    Usage:
        registry = NamedRangeRegistry()
        registry.define(wb, "WACC", "Assumptions", "B20")
        # In a formula on another sheet:
        f"=SomeCalc * {registry.ref('WACC')}"
        # For explicit cross-sheet:
        registry.sheet_ref("WACC")  ->  "'Assumptions'!$B$20"
    """

    def __init__(self) -> None:
        self._names: Dict[str, tuple[str, str]] = {}  # name -> (sheet_title, cell_ref)

    def define(self, wb: Workbook, name: str, sheet_title: str, cell_ref: str) -> str:
        """Create a workbook-scoped named range and store the mapping."""
        if "$" in cell_ref:
            abs_ref = cell_ref
        else:
            # Split letters (column) from digits (row) to handle multi-letter columns
            col_part = "".join(c for c in cell_ref if c.isalpha())
            row_part = "".join(c for c in cell_ref if c.isdigit())
            abs_ref = f"${col_part}${row_part}"
        self._names[name] = (sheet_title, abs_ref)
        formula_ref = f"'{sheet_title}'!{abs_ref}"
        try:
            defn = DefinedName(name, attr_text=formula_ref)
            wb.defined_names.add(defn)
        except Exception:
            logger.warning("Could not define named range %s (may already exist)", name)
        return name

    def ref(self, name: str) -> str:
        """Return the name itself, usable directly in Excel formulas."""
        return name

    def sheet_ref(self, name: str) -> str:
        """Return explicit 'SheetName'!$B$5 for cross-sheet formulas."""
        sheet_title, cell_ref = self._names[name]
        return f"'{sheet_title}'!{cell_ref}"

    def has(self, name: str) -> bool:
        return name in self._names


# =============================================================================
# SHARED HELPERS
# =============================================================================

def _apply_header_row(ws, row: int, values: List[str], start_col: int = 1) -> None:
    """Write a branded header row with consistent styling."""
    for col_offset, val in enumerate(values):
        cell = ws.cell(row=row, column=start_col + col_offset, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_label_value(
    ws, row: int, label: str, value,
    col_label: int = 1, col_value: int = 2,
    fmt: Optional[str] = None,
    font: Optional[Font] = None,
) -> None:
    """Write a label/value pair on a row."""
    lc = ws.cell(row=row, column=col_label, value=label)
    lc.font = LABEL_FONT
    lc.border = THIN_BORDER
    vc = ws.cell(row=row, column=col_value, value=value)
    vc.font = font or VALUE_FONT
    vc.border = THIN_BORDER
    vc.alignment = Alignment(horizontal="right")
    if fmt:
        vc.number_format = fmt


def _write_input_cell(ws, row: int, col: int, value, fmt: Optional[str] = None) -> None:
    """Write an editable blue input cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = INPUT_FONT
    cell.protection = UNLOCKED
    cell.alignment = Alignment(horizontal="right")
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt


def _write_formula_cell(
    ws, row: int, col: int, formula: str,
    fmt: Optional[str] = None,
    font: Optional[Font] = None,
    border: Optional[Border] = None,
) -> None:
    """Write a formula cell with appropriate styling."""
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = font or FORMULA_FONT
    cell.alignment = Alignment(horizontal="right")
    cell.border = border or THIN_BORDER
    if fmt:
        cell.number_format = fmt


def _auto_width(ws, min_width: int = 10, max_width: int = 45) -> None:
    """Auto-fit column widths based on content, wrapping text that exceeds max_width."""
    for col_cells in ws.columns:
        length = min_width
        for cell in col_cells:
            if cell.value is not None:
                content_len = len(str(cell.value)) + 3
                if content_len > max_width:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        horizontal=cell.alignment.horizontal or "left",
                        vertical=cell.alignment.vertical or "top",
                    )
                length = max(length, min(content_len, max_width))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length


def _safe_float(val, default=None) -> Optional[float]:
    """Safely convert a value to float."""
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _extract_year_label(stmt: Dict, is_estimate: bool = False) -> str:
    """Extract a readable fiscal year label with A (actual) or E (estimate) suffix.

    Suffix convention: FY2024A = historical actual, FY2025E = analyst estimate.
    Uses 'is_estimate' field from the statement dict if not explicitly passed.
    """
    est = is_estimate or stmt.get("is_estimate", False)
    suffix = "E" if est else "A"
    fy = stmt.get("fiscal_year")
    if fy and isinstance(fy, (int, float)) and fy > 0:
        return f"FY{int(fy)}{suffix}"
    period = str(stmt.get("period", ""))
    if period and len(period) >= 4:
        return f"FY{period[:4]}{suffix}"
    return "FY"


def _deduplicate_fiscal_years(statements: List[Dict]) -> List[Dict]:
    """Keep the last occurrence when two entries share a fiscal year label.

    Returns statements sorted by fiscal year ascending (oldest on the left).
    """
    seen: Dict[str, int] = {}
    for i, stmt in enumerate(statements):
        label = _extract_year_label(stmt)
        seen[label] = i
    indices = sorted(seen.values())
    deduped = [statements[i] for i in indices]
    # Sort ascending by fiscal year so oldest appears on the left
    deduped.sort(key=lambda s: s.get("fiscal_year", 0) or 0)
    return deduped


def _format_scenario_assumptions(assumptions) -> str:
    """Format scenario assumptions dict into readable prose for Excel."""
    if not assumptions:
        return ""
    if isinstance(assumptions, str):
        return assumptions[:500]
    if not isinstance(assumptions, dict):
        return str(assumptions)[:500]

    sentences = []

    # Lead with any explicit rationale from the LLM
    for key in ("revenue_growth_rationale", "fcf_margin_rationale", "margin_rationale"):
        rationale = assumptions.get(key, "")
        if rationale and "Conservative defaults" not in rationale:
            sentences.append(rationale.rstrip(".") + ".")
            break

    # Revenue growth trajectory
    rates = assumptions.get("revenue_growth_rates", [])
    if rates:
        first = rates[0] * 100
        last = rates[-1] * 100
        if len(rates) == 1:
            sentences.append(f"Revenue growth assumed at {first:.1f}%.")
        elif first > last:
            sentences.append(
                f"Revenue growth decelerates from {first:.1f}% to {last:.1f}% "
                f"over the {len(rates)}-year projection horizon."
            )
        elif first < last:
            sentences.append(
                f"Revenue growth accelerates from {first:.1f}% to {last:.1f}% "
                f"over the {len(rates)}-year projection horizon."
            )
        else:
            sentences.append(f"Revenue growth held constant at {first:.1f}%.")

    # WACC and TGR as a combined sentence
    wacc = assumptions.get("wacc")
    tgr = assumptions.get("terminal_growth_rate")
    if wacc is not None and tgr is not None:
        sentences.append(
            f"Discounted at a {wacc * 100:.1f}% WACC with a "
            f"{tgr * 100:.1f}% terminal growth rate."
        )
    elif wacc is not None:
        sentences.append(f"Discounted at a {wacc * 100:.1f}% WACC.")
    elif tgr is not None:
        sentences.append(f"Terminal growth rate of {tgr * 100:.1f}%.")

    # FCF margin if present
    fcf_margin = assumptions.get("fcf_margin")
    if fcf_margin is not None:
        sentences.append(f"FCF margin assumed at {fcf_margin * 100:.1f}%.")

    return " ".join(sentences)[:500]


def _col_letter(col: int) -> str:
    """1-indexed column number to letter (1='A', 2='B', ...)."""
    return get_column_letter(col)


def _cell_addr(col: int, row: int) -> str:
    """Return absolute cell address like $B$5."""
    return f"${_col_letter(col)}${row}"


# =============================================================================
# DEEP INTERNAL HELPERS
# =============================================================================

def _write_statement_section(
    ws, row: int, subtitle: str, statements: List[Dict],
    line_items: List[Tuple[str, str, str]],
    derived_ratios: Optional[List[Tuple[str, str, str, str]]] = None,
) -> Tuple[int, Dict[str, int]]:
    """Write a complete financial statement sub-section: subtitle, year headers,
    data rows with styling, and derived ratio formulas with division-by-zero guards.

    Args:
        line_items: [(label, dict_key, number_format), ...] for raw data rows.
        derived_ratios: [(label, numerator_key, denominator_key, number_format), ...]
            Each generates =IF(denom=0,0,num/denom) per year column.

    Returns:
        (next_row, item_row_map) where item_row_map maps dict_key -> row number.
    """
    n_stmts = min(len(statements), 5)
    item_rows: Dict[str, int] = {}

    ws.cell(row=row, column=1, value=subtitle).font = SUBTITLE_FONT
    row += 1

    ws.cell(row=row, column=1, value="").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    for j in range(n_stmts):
        cell = ws.cell(row=row, column=2 + j, value=_extract_year_label(statements[j]))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for label, key, fmt in line_items:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        item_rows[key] = row
        for j in range(n_stmts):
            val = _safe_float(statements[j].get(key))
            if val is not None:
                cell = ws.cell(row=row, column=2 + j, value=val)
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right")
                cell.border = THIN_BORDER
        row += 1

    if derived_ratios:
        for label, numerator_key, denominator_key, fmt in derived_ratios:
            num_row = item_rows.get(numerator_key)
            denom_row = item_rows.get(denominator_key)
            if num_row and denom_row:
                ws.cell(row=row, column=1, value=label).font = ITALIC_LABEL_FONT
                ws.cell(row=row, column=1).border = THIN_BORDER
                for j in range(n_stmts):
                    col = _col_letter(2 + j)
                    formula = f"=IF({col}{denom_row}=0,0,{col}{num_row}/{col}{denom_row})"
                    _write_formula_cell(ws, row, 2 + j, formula, fmt=fmt)
                row += 1

    return row, item_rows


def _write_projected_row(
    ws, row: int, label: str, n_years: int,
    formula_fn: Callable[[int, str], str],
    start_col: int = 3,
    fmt: Optional[str] = None,
    font: Optional[Font] = None,
    border: Optional[Border] = None,
    label_font: Optional[Font] = None,
    label_border: Optional[Border] = None,
) -> int:
    """Write one label + N formula cells across year columns.

    formula_fn(year_index, col_letter) returns the Excel formula string for
    that year's cell. Returns the row number written to (caller increments).
    """
    ws.cell(row=row, column=1, value=label).font = label_font or LABEL_FONT
    ws.cell(row=row, column=1).border = label_border or THIN_BORDER
    for i in range(n_years):
        col = _col_letter(start_col + i)
        _write_formula_cell(ws, row, start_col + i, formula_fn(i, col),
                            fmt=fmt, font=font, border=border)
    return row


def _write_year_input_row(
    ws, row: int, label: str, values: List,
    n_years: int, start_col: int,
    fmt: Optional[str],
    registry: "NamedRangeRegistry", wb: Workbook,
    sheet: str, name_prefix: str,
) -> int:
    """Write a label + N input cells with named range registration.

    Returns the row number written to (caller increments).
    """
    ws.cell(row=row, column=1, value=label).font = LABEL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    for i in range(n_years):
        v = _safe_float(values[i]) if i < len(values) else None
        _write_input_cell(ws, row, start_col + i, v, fmt=fmt)
        registry.define(wb, f"{name_prefix}{i + 1}", sheet,
                        f"{_col_letter(start_col + i)}{row}")
    return row


# =============================================================================
# ASSUMPTIONS SHEET
# =============================================================================

def _write_scenario_reference(
    ws, row: int, scenarios: Optional[Dict], base_assumptions: Dict,
) -> int:
    """Append a read-only Scenario Reference table after the EV Bridge section.

    Shows Bull/Base/Bear assumptions side-by-side when scenario dicts contain
    an 'assumptions' sub-dict. Returns the next available row. Skips entirely
    if no scenario carries assumption data (keeps test fixtures unchanged).
    """
    if not scenarios:
        return row

    # Collect scenario assumption dicts; skip if none exist
    case_keys = ("bull", "base", "bear")
    case_assumptions = {}
    for key in case_keys:
        case = scenarios.get(key, {})
        if isinstance(case, dict) and "assumptions" in case:
            case_assumptions[key] = case["assumptions"]
    if not case_assumptions:
        return row

    row += 1
    ws.cell(row=row, column=1, value="Scenario Reference").font = SUBTITLE_FONT
    row += 1

    # Header row
    headers = ["", "Bull", "Base", "Bear"]
    _apply_header_row(ws, row, headers)
    row += 1

    def _get_val(case_key: str, field: str, index: Optional[int] = None):
        a = case_assumptions.get(case_key, {})
        v = a.get(field)
        if index is not None and isinstance(v, (list, tuple)):
            return v[index] if index < len(v) else None
        return v

    # Rows to display: (label, field, index_or_None, fmt)
    metric_rows = [
        ("WACC", "wacc", None, FMT_PCT),
        ("Terminal Growth Rate", "terminal_growth_rate", None, FMT_PCT),
        ("Tax Rate", "tax_rate", None, FMT_PCT),
        ("Rev Growth Y1", "revenue_growth_rates", 0, FMT_PCT),
        ("Rev Growth Y3", "revenue_growth_rates", 2, FMT_PCT),
        ("Rev Growth Y5", "revenue_growth_rates", 4, FMT_PCT),
        ("Gross Margin Y1", "gross_margins", 0, FMT_PCT),
    ]

    for label, field, idx, fmt in metric_rows:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        for col_offset, case_key in enumerate(case_keys):
            val = _get_val(case_key, field, idx)
            if val is not None:
                val = _safe_float(val)
            if val is not None:
                cell = ws.cell(row=row, column=2 + col_offset, value=val)
                cell.font = LABEL_GRAY_FONT
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right")
                cell.border = THIN_BORDER
        row += 1

    ws.cell(row=row, column=1,
            value="Read-only reference. Edit assumptions in the sections above.").font = Font(
        name="Calibri", size=9, italic=True, color="808080",
    )
    return row + 1


def _build_assumptions_sheet(
    wb: Workbook,
    registry: NamedRangeRegistry,
    dcf: Optional[Dict],
    current_price: Optional[float],
    model_audit: Optional[Dict] = None,
    scenarios: Optional[Dict] = None,
) -> None:
    """Build the Assumptions sheet: single source of truth for all editable inputs.

    Returns nothing; all cell locations are registered via the NamedRangeRegistry
    so downstream sheets can reference them by name.
    """
    if not dcf:
        return

    ws = wb.create_sheet("Assumptions")
    sheet = "Assumptions"
    assumptions = dcf.get("assumptions", {})
    wacc_components = dcf.get("wacc_components")
    ev_bridge = dcf.get("ev_bridge", {})

    # Rationale map: row -> (col, text). Populated as rows are written, flushed at end.
    rationale_map: Dict[int, tuple[int, str]] = {}

    ws.cell(row=1, column=1, value="Model Assumptions").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Blue cells are editable inputs").font = LABEL_GRAY_FONT

    row = 4

    # --- Company Information ---
    ws.cell(row=row, column=1, value="Company Information").font = SUBTITLE_FONT
    row += 1

    base_rev = _safe_float(dcf.get("base_revenue"))
    shares = _safe_float(dcf.get("shares_outstanding"))
    price = _safe_float(current_price)

    for label, val, fmt, name in [
        ("Base Revenue", base_rev, FMT_BILLIONS, "BaseRevenue"),
        ("Shares Outstanding", shares, FMT_NUMBER, "SharesOutstanding"),
        ("Current Price", price, FMT_USD, "CurrentPrice"),
    ]:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        _write_input_cell(ws, row, 2, val, fmt=fmt)
        registry.define(wb, name, sheet, f"B{row}")
        row += 1

    row += 1

    # --- WACC Build-Up ---
    ws.cell(row=row, column=1, value="WACC Build-Up").font = SUBTITLE_FONT
    row += 1

    if wacc_components:
        # Full decomposed WACC with formulas
        wc = wacc_components
        input_fields = [
            ("Risk-Free Rate", _safe_float(wc.get("risk_free_rate")), FMT_PCT, "Rf"),
            ("Beta", _safe_float(wc.get("beta")), "0.00", "Beta"),
            ("Equity Risk Premium", _safe_float(wc.get("equity_risk_premium")), FMT_PCT, "ERP"),
            ("Size Premium", _safe_float(wc.get("size_premium")), FMT_PCT, "SizePremium"),
        ]
        for label, val, fmt, name in input_fields:
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            _write_input_cell(ws, row, 2, val, fmt=fmt)
            registry.define(wb, name, sheet, f"B{row}")
            row += 1

        # Ke = Rf + Beta*ERP + SP (formula)
        ws.cell(row=row, column=1, value="Cost of Equity (Ke)").font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ke_formula = f"={registry.ref('Rf')}+{registry.ref('Beta')}*{registry.ref('ERP')}+{registry.ref('SizePremium')}"
        _write_formula_cell(ws, row, 2, ke_formula, fmt=FMT_PCT)
        registry.define(wb, "Ke", sheet, f"B{row}")
        row += 1

        # Kd pre-tax (input)
        ws.cell(row=row, column=1, value="Cost of Debt (pre-tax)").font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        _write_input_cell(ws, row, 2, _safe_float(wc.get("kd_pretax")), fmt=FMT_PCT)
        registry.define(wb, "KdPreTax", sheet, f"B{row}")
        row += 1

        # Tax Rate (input)
        ws.cell(row=row, column=1, value="Tax Rate").font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        _write_input_cell(ws, row, 2, _safe_float(wc.get("tax_rate")), fmt=FMT_PCT)
        registry.define(wb, "TaxRate", sheet, f"B{row}")
        row += 1

        # Kd after-tax = KdPreTax * (1 - TaxRate)
        ws.cell(row=row, column=1, value="Cost of Debt (after-tax)").font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        kd_formula = f"={registry.ref('KdPreTax')}*(1-{registry.ref('TaxRate')})"
        _write_formula_cell(ws, row, 2, kd_formula, fmt=FMT_PCT)
        registry.define(wb, "KdAfterTax", sheet, f"B{row}")
        row += 1

        # Weight Equity (input)
        ws.cell(row=row, column=1, value="Weight Equity").font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        _write_input_cell(ws, row, 2, _safe_float(wc.get("weight_equity")), fmt=FMT_PCT)
        registry.define(wb, "WeightEquity", sheet, f"B{row}")
        row += 1

        # Weight Debt = 1 - WeightEquity
        ws.cell(row=row, column=1, value="Weight Debt").font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        wd_formula = f"=1-{registry.ref('WeightEquity')}"
        _write_formula_cell(ws, row, 2, wd_formula, fmt=FMT_PCT)
        registry.define(wb, "WeightDebt", sheet, f"B{row}")
        row += 1

        # WACC = Ke*We + Kd_at*Wd
        ws.cell(row=row, column=1, value="WACC").font = LABEL_FONT
        ws.cell(row=row, column=1).border = TOTAL_BORDER
        wacc_formula = (
            f"={registry.ref('Ke')}*{registry.ref('WeightEquity')}"
            f"+{registry.ref('KdAfterTax')}*(1-{registry.ref('WeightEquity')})"
        )
        _write_formula_cell(ws, row, 2, wacc_formula, fmt=FMT_PCT, border=TOTAL_BORDER)
        registry.define(wb, "WACC", sheet, f"B{row}")
        wacc_rationale = assumptions.get("wacc_rationale", "")
        if not wacc_rationale and model_audit:
            wacc_rationale = (model_audit.get("assumption_notes") or {}).get("wacc", "")
        if wacc_rationale:
            rationale_map[row] = (3, wacc_rationale)
        row += 1
    else:
        # Fallback: single WACC input cell
        wacc_val = _safe_float(assumptions.get("wacc"))
        ws.cell(row=row, column=1, value="WACC").font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        _write_input_cell(ws, row, 2, wacc_val, fmt=FMT_PCT)
        registry.define(wb, "WACC", sheet, f"B{row}")
        row += 1

        # Still need TaxRate for DCF sheet
        tax = _safe_float(assumptions.get("tax_rate", 0.21))
        ws.cell(row=row, column=1, value="Tax Rate").font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        _write_input_cell(ws, row, 2, tax, fmt=FMT_PCT)
        registry.define(wb, "TaxRate", sheet, f"B{row}")
        row += 1

    row += 1

    # --- Terminal Value ---
    ws.cell(row=row, column=1, value="Terminal Value").font = SUBTITLE_FONT
    row += 1
    tgr = _safe_float(assumptions.get("terminal_growth_rate"))
    ws.cell(row=row, column=1, value="Terminal Growth Rate").font = LABEL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    _write_input_cell(ws, row, 2, tgr, fmt=FMT_PCT)
    registry.define(wb, "TGR", sheet, f"B{row}")
    tgr_rationale = assumptions.get("terminal_growth_rationale", "")
    if not tgr_rationale and model_audit:
        tgr_rationale = (model_audit.get("assumption_notes") or {}).get("terminal_growth", "")
    if tgr_rationale:
        rationale_map[row] = (3, tgr_rationale)
    row += 1

    # Mid-year convention toggle (1 = on, 0 = off)
    ws.cell(row=row, column=1, value="Mid-Year Convention").font = LABEL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    _write_input_cell(ws, row, 2, 0, fmt="0")
    registry.define(wb, "MidYearConvention", sheet, f"B{row}")
    row += 1

    row += 1

    # --- Projection Assumptions (year-by-year) ---
    growth_rates = assumptions.get("revenue_growth_rates", [])
    n_years = len(growth_rates) if growth_rates else 5

    ws.cell(row=row, column=1, value="Projection Assumptions").font = SUBTITLE_FONT
    row += 1

    # Year headers
    ws.cell(row=row, column=1, value="").font = HEADER_FONT
    for i in range(n_years):
        cell = ws.cell(row=row, column=2 + i, value=f"Y{i + 1}")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Revenue Growth Rates
    _write_year_input_row(ws, row, "Revenue Growth", growth_rates,
                          n_years, 2, FMT_PCT, registry, wb, sheet, "GrowthY")
    rev_rationale = assumptions.get("revenue_growth_rationale", "")
    if rev_rationale:
        rationale_map[row] = (n_years + 2, rev_rationale)
    row += 1

    # Decomposed margin assumptions (if available)
    gross_margins = assumptions.get("gross_margins")
    is_decomposed = gross_margins is not None and len(gross_margins) > 0

    if is_decomposed:
        margin_rationale = assumptions.get("margin_rationale", "")
        for label, values, prefix in [
            ("Gross Margin", gross_margins, "GrossMarginY"),
            ("OpEx % Revenue", assumptions.get("opex_as_pct_revenue", []), "OpExPctY"),
            ("CapEx % Revenue", assumptions.get("capex_as_pct_revenue", []), "CapExPctY"),
        ]:
            _write_year_input_row(ws, row, label, values or [],
                                  n_years, 2, FMT_PCT, registry, wb, sheet, prefix)
            # Attach margin rationale to Gross Margin and OpEx rows
            if margin_rationale and prefix in ("GrossMarginY", "OpExPctY"):
                rationale_map[row] = (n_years + 2, margin_rationale)
            row += 1

        # Per-year D&A, SBC, NWC: scalar expanded to N years so the analyst
        # can vary them year by year while the default is uniform.
        for label, scalar_val, prefix in [
            ("D&A % Revenue", assumptions.get("da_as_pct_revenue"), "DAPctY"),
            ("SBC % Revenue", assumptions.get("sbc_as_pct_revenue"), "SBCPctY"),
            ("NWC % Rev Delta", assumptions.get("nwc_change_pct"), "NWCPctY"),
        ]:
            sv = _safe_float(scalar_val)
            _write_year_input_row(ws, row, label, [sv] * n_years,
                                  n_years, 2, FMT_PCT, registry, wb, sheet, prefix)
            row += 1
    else:
        # Legacy: single FCF margin
        fcf_margin = _safe_float(assumptions.get("fcf_margin"))
        ws.cell(row=row, column=1, value="FCF Margin").font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        _write_input_cell(ws, row, 2, fcf_margin, fmt=FMT_PCT)
        registry.define(wb, "FCFMargin", sheet, f"B{row}")
        row += 1

    row += 1

    # --- EV Bridge Components ---
    ws.cell(row=row, column=1, value="EV Bridge").font = SUBTITLE_FONT
    row += 1

    net_debt = _safe_float(ev_bridge.get("net_debt") if ev_bridge else dcf.get("net_debt"))
    bridge_items = [
        ("Net Debt", net_debt, "NetDebt"),
    ]
    # Only show nonzero EV bridge items
    for label, key, name in [
        ("Minority Interest", "minority_interest", "MinorityInterest"),
        ("Preferred Stock", "preferred_stock", "PreferredStock"),
        ("Pension Deficit", "pension_deficit", "PensionDeficit"),
        ("Equity Investments", "equity_investments", "EquityInvestments"),
    ]:
        val = _safe_float(ev_bridge.get(key)) if ev_bridge else None
        if val and val != 0:
            bridge_items.append((label, val, name))

    for label, val, name in bridge_items:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        _write_input_cell(ws, row, 2, val, fmt=FMT_BILLIONS)
        registry.define(wb, name, sheet, f"B{row}")
        row += 1

    # --- Scenario Reference Table (Feature 2) ---
    _write_scenario_reference(ws, row, scenarios, assumptions)

    # --- Flush rationale cells ---
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for r, (c, text) in rationale_map.items():
        cell = ws.cell(row=r, column=c, value=text)
        cell.font = RATIONALE_FONT
        cell.alignment = wrap_align

    _auto_width(ws)


# =============================================================================
# DCF SHEET
# =============================================================================

def _build_dcf_sheet(
    wb: Workbook,
    registry: NamedRangeRegistry,
    dcf: Dict,
) -> None:
    """Build DCF sheet with full formula chain referencing Assumptions."""
    ws = wb.create_sheet("DCF Model")
    assumptions = dcf.get("assumptions", {})
    has_assumptions_sheet = registry.has("WACC")

    ws.cell(row=1, column=1, value="Discounted Cash Flow Analysis").font = TITLE_FONT
    ws.cell(row=2, column=1, value="(All figures in USD billions unless otherwise stated)").font = LABEL_GRAY_FONT

    growth_rates = assumptions.get("revenue_growth_rates", [])
    n_years = len(growth_rates) if growth_rates else 5
    gross_margins = assumptions.get("gross_margins")
    is_decomposed = gross_margins is not None and len(gross_margins) > 0

    # Year headers
    row = 3
    ws.cell(row=row, column=1, value="").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    base_cell = ws.cell(row=row, column=2, value="Base")
    base_cell.font = HEADER_FONT
    base_cell.fill = HEADER_FILL
    base_cell.alignment = Alignment(horizontal="center")
    for i in range(n_years):
        cell = ws.cell(row=row, column=3 + i, value=f"Y{i + 1}")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1

    if not has_assumptions_sheet:
        # Fallback: write static values (old behavior)
        _build_dcf_sheet_static(ws, dcf, row)
        return

    # --- Revenue row (special: also writes base column) ---
    rev_row = row
    _write_projected_row(ws, row, "Revenue", n_years,
        lambda i, col: f"={_col_letter(2 + i)}{rev_row}*(1+{registry.sheet_ref(f'GrowthY{i+1}')})",
        fmt=FMT_BILLIONS, font=LINK_FONT)
    _write_formula_cell(ws, row, 2, f"={registry.sheet_ref('BaseRevenue')}",
                        fmt=FMT_BILLIONS, font=LINK_FONT)
    row += 1

    if is_decomposed:
        gp_row = _write_projected_row(ws, row, "Gross Profit", n_years,
            lambda i, col: f"={col}{rev_row}*{registry.sheet_ref(f'GrossMarginY{i+1}')}",
            fmt=FMT_BILLIONS, font=LINK_FONT)
        row = gp_row + 1

        opex_row = _write_projected_row(ws, row, "(-) OpEx", n_years,
            lambda i, col: f"={col}{rev_row}*{registry.sheet_ref(f'OpExPctY{i+1}')}",
            fmt=FMT_BILLIONS, font=LINK_FONT)
        row = opex_row + 1

        ebitda_row = _write_projected_row(ws, row, "EBITDA", n_years,
            lambda i, col: f"={col}{gp_row}-{col}{opex_row}",
            fmt=FMT_BILLIONS, border=SUBTOTAL_BORDER, label_border=SUBTOTAL_BORDER)
        row = ebitda_row + 1

        da_row = _write_projected_row(ws, row, "(-) D&A", n_years,
            lambda i, col: f"={col}{rev_row}*{registry.sheet_ref(f'DAPctY{i+1}')}",
            fmt=FMT_BILLIONS, font=LINK_FONT)
        row = da_row + 1

        ebit_row = _write_projected_row(ws, row, "EBIT", n_years,
            lambda i, col: f"={col}{ebitda_row}-{col}{da_row}",
            fmt=FMT_BILLIONS)
        row = ebit_row + 1

        tax_ref = registry.sheet_ref("TaxRate")
        tax_row = _write_projected_row(ws, row, "(-) Tax", n_years,
            lambda i, col: f"={col}{ebit_row}*{tax_ref}",
            fmt=FMT_BILLIONS, font=LINK_FONT)
        row = tax_row + 1

        nopat_row = _write_projected_row(ws, row, "NOPAT", n_years,
            lambda i, col: f"={col}{ebit_row}-{col}{tax_row}",
            fmt=FMT_BILLIONS, border=SUBTOTAL_BORDER, label_border=SUBTOTAL_BORDER)
        row = nopat_row + 1

        da_addback_row = _write_projected_row(ws, row, "(+) D&A Add-back", n_years,
            lambda i, col: f"={col}{da_row}",
            fmt=FMT_BILLIONS)
        row = da_addback_row + 1

        capex_row = _write_projected_row(ws, row, "(-) CapEx", n_years,
            lambda i, col: f"={col}{rev_row}*{registry.sheet_ref(f'CapExPctY{i+1}')}",
            fmt=FMT_BILLIONS, font=LINK_FONT)
        row = capex_row + 1

        nwc_row = _write_projected_row(ws, row, "(-) NWC Change", n_years,
            lambda i, col: (
                f"=({col}{rev_row}-{_col_letter(2 + i)}{rev_row})"
                f"*{registry.sheet_ref(f'NWCPctY{i+1}')}"
            ),
            fmt=FMT_BILLIONS, font=LINK_FONT)
        row = nwc_row + 1

        sbc_row = _write_projected_row(ws, row, "(+) SBC Add-back", n_years,
            lambda i, col: f"={col}{rev_row}*{registry.sheet_ref(f'SBCPctY{i+1}')}",
            fmt=FMT_BILLIONS, font=LINK_FONT)
        row = sbc_row + 1

        fcf_row = _write_projected_row(ws, row, "Free Cash Flow", n_years,
            lambda i, col: (
                f"={col}{nopat_row}+{col}{da_addback_row}"
                f"-{col}{capex_row}-{col}{nwc_row}+{col}{sbc_row}"
            ),
            fmt=FMT_BILLIONS, border=TOTAL_BORDER,
            label_font=Font(name="Calibri", bold=True, size=11),
            label_border=TOTAL_BORDER)
        row = fcf_row + 1

    else:
        # Legacy mode: FCF = Revenue * Margin
        margin_ref = registry.sheet_ref("FCFMargin")
        _write_projected_row(ws, row, "FCF Margin", n_years,
            lambda i, col: f"={margin_ref}",
            fmt=FMT_PCT, font=LINK_FONT)
        row += 1

        fcf_row = _write_projected_row(ws, row, "Free Cash Flow", n_years,
            lambda i, col: f"={col}{rev_row}*{margin_ref}",
            fmt=FMT_BILLIONS, font=LINK_FONT, border=TOTAL_BORDER,
            label_font=Font(name="Calibri", bold=True, size=11),
            label_border=TOTAL_BORDER)
        row = fcf_row + 1

    row += 1

    # --- Discount Factor & PV of FCF ---
    wacc_ref = registry.sheet_ref("WACC")
    df_row = _write_projected_row(ws, row, "Discount Factor", n_years,
        lambda i, col: f"=1/(1+{wacc_ref})^({i + 1}-0.5*MidYearConvention)",
        fmt="0.0000", font=LINK_FONT)
    row = df_row + 1

    pv_fcf_row = _write_projected_row(ws, row, "PV of FCF", n_years,
        lambda i, col: f"={col}{fcf_row}*{col}{df_row}",
        fmt=FMT_BILLIONS)
    row = pv_fcf_row + 2

    # --- Terminal Value ---
    ws.cell(row=row, column=1, value="Terminal Value").font = SUBTITLE_FONT
    row += 1

    # Final year FCF
    ws.cell(row=row, column=1, value="Final Year FCF").font = LABEL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    final_fcf_row = row
    last_col = _col_letter(2 + n_years)
    _write_formula_cell(ws, row, 2, f"={last_col}{fcf_row}", fmt=FMT_BILLIONS)
    row += 1

    # TV = FCF*(1+g)/(WACC-g)
    ws.cell(row=row, column=1, value="Terminal Value").font = LABEL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    tv_row = row
    tgr_ref = registry.sheet_ref("TGR")
    tv_formula = f"=B{final_fcf_row}*(1+{tgr_ref})/({wacc_ref}-{tgr_ref})"
    _write_formula_cell(ws, row, 2, tv_formula, fmt=FMT_BILLIONS)
    registry.define(wb, "TerminalValue", "DCF Model", f"B{row}")
    row += 1

    # PV of TV
    ws.cell(row=row, column=1, value="PV of Terminal Value").font = LABEL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    pv_tv_row = row
    _write_formula_cell(ws, row, 2, f"=B{tv_row}/(1+{wacc_ref})^{n_years}", fmt=FMT_BILLIONS)
    registry.define(wb, "PVTerminalValue", "DCF Model", f"B{row}")
    row += 2

    # --- EV-to-Equity Bridge ---
    ws.cell(row=row, column=1, value="Enterprise Value to Equity Bridge").font = SUBTITLE_FONT
    row += 1

    # PV of FCFs = SUM
    ws.cell(row=row, column=1, value="PV of FCFs").font = LABEL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    pv_fcfs_row = row
    first_pv = f"C{pv_fcf_row}"
    last_pv = f"{_col_letter(2 + n_years)}{pv_fcf_row}"
    _write_formula_cell(ws, row, 2, f"=SUM({first_pv}:{last_pv})", fmt=FMT_BILLIONS)
    row += 1

    # (+) PV of Terminal
    ws.cell(row=row, column=1, value="(+) PV of Terminal Value").font = LABEL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    _write_formula_cell(ws, row, 2, f"=B{pv_tv_row}", fmt=FMT_BILLIONS)
    pv_tv_bridge_row = row
    row += 1

    # Enterprise Value
    ws.cell(row=row, column=1, value="Enterprise Value").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=row, column=1).border = SUBTOTAL_BORDER
    ev_row = row
    _write_formula_cell(ws, row, 2, f"=B{pv_fcfs_row}+B{pv_tv_bridge_row}",
                        fmt=FMT_BILLIONS, border=SUBTOTAL_BORDER)
    registry.define(wb, "EnterpriseValue", "DCF Model", f"B{row}")
    row += 1

    # EV bridge deductions
    deduction_refs = []
    addition_refs = []

    net_debt_ref = registry.sheet_ref("NetDebt")
    ws.cell(row=row, column=1, value="(-) Net Debt").font = LABEL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    _write_formula_cell(ws, row, 2, f"={net_debt_ref}", fmt=FMT_BILLIONS, font=LINK_FONT)
    deduction_refs.append(f"B{row}")
    row += 1

    for label, name, is_addition in [
        ("(-) Minority Interest", "MinorityInterest", False),
        ("(-) Preferred Stock", "PreferredStock", False),
        ("(-) Pension Deficit", "PensionDeficit", False),
        ("(+) Equity Investments", "EquityInvestments", True),
    ]:
        if registry.has(name):
            ref = registry.sheet_ref(name)
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            _write_formula_cell(ws, row, 2, f"={ref}", fmt=FMT_BILLIONS, font=LINK_FONT)
            if is_addition:
                addition_refs.append(f"B{row}")
            else:
                deduction_refs.append(f"B{row}")
            row += 1

    # Equity Value = EV - deductions + additions
    ws.cell(row=row, column=1, value="Equity Value").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=row, column=1).border = TOTAL_BORDER
    eq_formula = f"=B{ev_row}"
    for ref in deduction_refs:
        eq_formula += f"-{ref}"
    for ref in addition_refs:
        eq_formula += f"+{ref}"
    _write_formula_cell(ws, row, 2, eq_formula, fmt=FMT_BILLIONS, border=TOTAL_BORDER)
    equity_val_row = row
    registry.define(wb, "EquityValue", "DCF Model", f"B{row}")
    row += 1

    # Shares Outstanding
    ws.cell(row=row, column=1, value="Shares Outstanding").font = LABEL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    shares_ref = registry.sheet_ref("SharesOutstanding")
    _write_formula_cell(ws, row, 2, f"={shares_ref}", fmt=FMT_NUMBER, font=LINK_FONT)
    shares_row = row
    row += 1

    # Fair Value / Share
    ws.cell(row=row, column=1, value="Fair Value / Share").font = Font(name="Calibri", bold=True, size=12)
    ws.cell(row=row, column=1).border = TOTAL_BORDER
    _write_formula_cell(ws, row, 2, f"=B{equity_val_row}/B{shares_row}",
                        fmt=FMT_USD, border=TOTAL_BORDER)
    registry.define(wb, "FairValue", "DCF Model", f"B{row}")
    fv_row = row
    row += 1

    # Upside / Downside
    if registry.has("CurrentPrice"):
        ws.cell(row=row, column=1, value="Upside / Downside").font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        price_ref = registry.sheet_ref("CurrentPrice")
        _write_formula_cell(ws, row, 2, f"=(B{fv_row}-{price_ref})/{price_ref}", fmt=FMT_PCT, font=LINK_FONT)
        row += 1

    _auto_width(ws)


def _build_dcf_sheet_static(ws, dcf: Dict, start_row: int) -> None:
    """Fallback: write static DCF values when no Assumptions sheet exists."""
    row = start_row
    assumptions = dcf.get("assumptions", {})

    # Key assumptions as static values
    for label, val, fmt in [
        ("WACC", assumptions.get("wacc"), FMT_PCT),
        ("Terminal Growth Rate", assumptions.get("terminal_growth_rate"), FMT_PCT),
        ("Tax Rate", assumptions.get("tax_rate"), FMT_PCT),
    ]:
        if val is not None:
            _write_label_value(ws, row, label, _safe_float(val), fmt=fmt)
            row += 1

    growth_rates = assumptions.get("revenue_growth_rates", [])
    if growth_rates:
        _write_label_value(ws, row, "Revenue Growth Rates",
                           ", ".join(f"{g * 100:.1f}%" for g in growth_rates))
        row += 1

    revenues = dcf.get("projected_revenues", [])
    fcfs = dcf.get("projected_fcfs", [])
    disc_fcfs = dcf.get("discounted_fcfs", [])

    if revenues or fcfs:
        row += 1
        n_years = max(len(revenues), len(fcfs), len(disc_fcfs))
        headers = ["Year"] + [f"Y{i + 1}" for i in range(n_years)]
        _apply_header_row(ws, row, headers)
        row += 1

        for label, data in [("Revenue", revenues), ("Free Cash Flow", fcfs), ("Discounted FCF", disc_fcfs)]:
            if data:
                ws.cell(row=row, column=1, value=label).font = LABEL_FONT
                for i, v in enumerate(data):
                    ws.cell(row=row, column=2 + i, value=_safe_float(v)).number_format = FMT_BILLIONS
                row += 1

    row += 1
    bridge_items = [
        ("PV of FCFs", dcf.get("pv_fcfs"), FMT_BILLIONS),
        ("Terminal Value", dcf.get("terminal_value"), FMT_BILLIONS),
        ("PV of Terminal Value", dcf.get("pv_terminal_value"), FMT_BILLIONS),
        ("Enterprise Value", dcf.get("enterprise_value"), FMT_BILLIONS),
        ("Net Debt", dcf.get("net_debt"), FMT_BILLIONS),
        ("Equity Value", dcf.get("equity_value"), FMT_BILLIONS),
        ("Shares Outstanding", dcf.get("shares_outstanding"), FMT_NUMBER),
        ("Fair Value / Share", dcf.get("fair_value_per_share"), FMT_USD),
    ]
    for label, val, fmt in bridge_items:
        fval = _safe_float(val)
        if fval is not None:
            _write_label_value(ws, row, label, fval, fmt=fmt)
            row += 1

    _auto_width(ws)


# =============================================================================
# FINANCIAL STATEMENTS SHEET
# =============================================================================

def _build_financial_statements_sheet(wb: Workbook, financials: Dict) -> None:
    """Build financial statements with derived ratio formulas.

    Historical line items are pasted values. Margins and ratios become
    formulas dividing one line item by another.
    """
    ws = wb.create_sheet("Financials")
    ws.cell(row=1, column=1, value="Financial Statements").font = TITLE_FONT
    ws.cell(row=2, column=1, value="(All figures in USD billions unless otherwise stated)").font = LABEL_GRAY_FONT

    row = 3
    is_item_rows: Dict[str, int] = {}
    n_is_stmts = 0

    # --- Income Statement ---
    stmts = _deduplicate_fiscal_years(financials.get("income_statements", []))
    if stmts:
        row, is_item_rows = _write_statement_section(
            ws, row, "Income Statement", stmts,
            line_items=[
                ("Revenue", "revenue", FMT_BILLIONS),
                ("Gross Profit", "gross_profit", FMT_BILLIONS),
                ("Operating Income", "operating_income", FMT_BILLIONS),
                ("Net Income", "net_income", FMT_BILLIONS),
                ("EPS (Diluted)", "eps_diluted", FMT_USD_INNER),
            ],
            derived_ratios=[
                ("Gross Margin", "gross_profit", "revenue", FMT_PCT),
                ("Operating Margin", "operating_income", "revenue", FMT_PCT),
                ("Net Margin", "net_income", "revenue", FMT_PCT),
            ],
        )
        n_is_stmts = min(len(stmts), 5)
        row += 1

    # --- Balance Sheet ---
    sheets = _deduplicate_fiscal_years(financials.get("balance_sheets", []))
    if sheets:
        row, bs_rows = _write_statement_section(
            ws, row, "Balance Sheet", sheets,
            line_items=[
                ("Total Assets", "total_assets", FMT_BILLIONS),
                ("Total Liabilities", "total_liabilities", FMT_BILLIONS),
                ("Total Equity", "total_equity", FMT_BILLIONS),
                ("Cash & Equivalents", "cash", FMT_BILLIONS),
                ("Total Debt", "total_debt", FMT_BILLIONS),
            ],
            derived_ratios=[
                ("Debt / Equity", "total_debt", "total_equity", FMT_RATIO),
            ],
        )

        # Current ratio is static data (no current assets/liabilities breakdown)
        n_sheets = min(len(sheets), 5)
        ws.cell(row=row, column=1, value="Current Ratio").font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        for j in range(n_sheets):
            val = _safe_float(sheets[j].get("current_ratio"))
            if val is not None:
                cell = ws.cell(row=row, column=2 + j, value=val)
                cell.number_format = FMT_RATIO
                cell.alignment = Alignment(horizontal="right")
                cell.border = THIN_BORDER
        row += 1

        row += 1

    # --- Cash Flow Statement ---
    cfs = _deduplicate_fiscal_years(financials.get("cash_flows", []))
    if cfs:
        row, cf_rows = _write_statement_section(
            ws, row, "Cash Flow Statement", cfs,
            line_items=[
                ("Operating Cash Flow", "operating_cf", FMT_BILLIONS),
                ("Investing Cash Flow", "investing_cf", FMT_BILLIONS),
                ("Financing Cash Flow", "financing_cf", FMT_BILLIONS),
                ("Free Cash Flow", "free_cash_flow", FMT_BILLIONS),
                ("Capex", "capex", FMT_BILLIONS),
            ],
        )

        # FCF Margin: cross-ref IS revenue when year counts match, else static
        n_cfs = min(len(cfs), 5)
        fcf_row_pos = cf_rows.get("free_cash_flow")
        rev_row_for_cf = (
            is_item_rows.get("revenue")
            if stmts and n_is_stmts == n_cfs
            else None
        )

        if fcf_row_pos:
            ws.cell(row=row, column=1, value="FCF Margin").font = ITALIC_LABEL_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            for j in range(n_cfs):
                col = _col_letter(2 + j)
                if rev_row_for_cf:
                    formula = f"=IF({col}{rev_row_for_cf}=0,0,{col}{fcf_row_pos}/{col}{rev_row_for_cf})"
                    _write_formula_cell(ws, row, 2 + j, formula, fmt=FMT_PCT, font=LINK_FONT)
                else:
                    val = _safe_float(cfs[j].get("fcf_margin"))
                    if val is not None:
                        cell = ws.cell(row=row, column=2 + j, value=val)
                        cell.number_format = FMT_PCT
                        cell.alignment = Alignment(horizontal="right")
                        cell.border = THIN_BORDER
            row += 1

    _auto_width(ws)


# =============================================================================
# SENSITIVITY SHEET (pre-computed grids)
# =============================================================================

def _build_sensitivity_sheet(
    wb: Workbook,
    sensitivity: Optional[Dict],
    growth_margin: Optional[Dict],
    current_price: Optional[float] = None,
) -> None:
    ws = wb.create_sheet("Sensitivity")
    ws.cell(row=1, column=1, value="Sensitivity Analysis").font = TITLE_FONT
    ws.cell(row=2, column=1, value="(Fair Value per Share in USD)").font = LABEL_GRAY_FONT

    row = 4

    if sensitivity:
        row = _write_sensitivity_grid(
            ws, row, "WACC vs Terminal Growth Rate",
            sensitivity.get("wacc_values", []),
            sensitivity.get("terminal_growth_values", []),
            sensitivity.get("fair_value_grid", []),
            sensitivity.get("base_wacc_idx"),
            sensitivity.get("base_tgr_idx"),
            row_label="WACC", col_label="TGR",
            current_price=current_price,
        )
        row += 2

    if growth_margin:
        row = _write_sensitivity_grid(
            ws, row, "Revenue Growth vs Margin",
            growth_margin.get("growth_values", growth_margin.get("row_values", [])),
            growth_margin.get("margin_values", growth_margin.get("col_values", [])),
            growth_margin.get("fair_value_grid", []),
            growth_margin.get("base_growth_idx", growth_margin.get("base_row_idx")),
            growth_margin.get("base_margin_idx", growth_margin.get("base_col_idx")),
            row_label="Growth", col_label="Margin",
            current_price=current_price,
        )

    _auto_width(ws)


def _write_sensitivity_grid(
    ws, start_row: int, title: str,
    row_values: List, col_values: List, grid: List[List],
    base_row_idx: Optional[int], base_col_idx: Optional[int],
    row_label: str = "Row", col_label: str = "Col",
    current_price: Optional[float] = None,
) -> int:
    """Write a sensitivity grid and return the next available row."""
    ws.cell(row=start_row, column=1, value=title).font = SUBTITLE_FONT
    row = start_row + 1

    ws.cell(row=row, column=1, value=f"FV/Share \\ {col_label}").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    for j, cv in enumerate(col_values):
        cell = ws.cell(row=row, column=2 + j, value=_safe_float(cv))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = FMT_PCT
    row += 1

    highlight_fill = PatternFill(
        start_color=_hex_strip(_PALETTE.pdf.base_case_bg),
        end_color=_hex_strip(_PALETTE.pdf.base_case_bg), fill_type="solid",
    )
    highlight_font = Font(name="Calibri", bold=True, size=11)
    above_price_fill = PatternFill(start_color=SUCCESS_COLOR, end_color=SUCCESS_COLOR, fill_type="solid")
    below_price_fill = PatternFill(start_color=DANGER_COLOR, end_color=DANGER_COLOR, fill_type="solid")

    for i, rv in enumerate(row_values):
        rh = ws.cell(row=row, column=1, value=_safe_float(rv))
        rh.font = LABEL_FONT
        rh.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        rh.number_format = FMT_PCT

        grid_row = grid[i] if i < len(grid) else []
        for j, val in enumerate(grid_row):
            cell = ws.cell(row=row, column=2 + j, value=_safe_float(val))
            cell.number_format = FMT_USD_INNER
            cell.alignment = Alignment(horizontal="center")

            is_base = (base_row_idx is not None and base_col_idx is not None
                       and i == base_row_idx and j == base_col_idx)
            if is_base:
                cell.fill = highlight_fill
                cell.font = highlight_font
            elif current_price is not None and val is not None:
                fval = _safe_float(val)
                if fval is not None:
                    cell.fill = above_price_fill if fval >= current_price else below_price_fill
        row += 1

    # Implied Upside / Downside companion grid
    if current_price and current_price > 0:
        row += 1
        ws.cell(row=row, column=1, value=f"{title} — Implied Upside / Downside").font = SUBTITLE_FONT
        row += 1

        ws.cell(row=row, column=1, value=f"{row_label} \\ {col_label}").font = HEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        for j, cv in enumerate(col_values):
            cell = ws.cell(row=row, column=2 + j, value=_safe_float(cv))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = FMT_PCT
        row += 1

        for i, rv in enumerate(row_values):
            rh = ws.cell(row=row, column=1, value=_safe_float(rv))
            rh.font = LABEL_FONT
            rh.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            rh.number_format = FMT_PCT

            grid_row = grid[i] if i < len(grid) else []
            for j, val in enumerate(grid_row):
                fval = _safe_float(val)
                if fval is not None and fval > 0:
                    upside = (fval - current_price) / current_price
                    cell = ws.cell(row=row, column=2 + j, value=upside)
                    cell.number_format = FMT_PCT
                    cell.alignment = Alignment(horizontal="center")

                    is_base = (base_row_idx is not None and base_col_idx is not None
                               and i == base_row_idx and j == base_col_idx)
                    if is_base:
                        cell.fill = highlight_fill
                        cell.font = highlight_font
                    elif upside >= 0:
                        cell.fill = above_price_fill
                    else:
                        cell.fill = below_price_fill
                else:
                    ws.cell(row=row, column=2 + j, value="N/A").alignment = Alignment(horizontal="center")
            row += 1

    return row


# =============================================================================
# SCENARIOS SHEET
# =============================================================================

def _build_scenarios_sheet(
    wb: Workbook,
    registry: NamedRangeRegistry,
    scenarios: Dict,
    current_price: Optional[float],
) -> None:
    ws = wb.create_sheet("Scenarios")
    ws.cell(row=1, column=1, value="Scenario Analysis").font = TITLE_FONT

    row = 3
    _apply_header_row(ws, row, ["Scenario", "Probability", "Fair Value", "Upside/Downside", "Key Assumptions"])
    row += 1

    scenario_rows: Dict[str, Dict[str, int]] = {}  # label -> {"prob": row, "fv": row}
    has_price = registry.has("CurrentPrice")

    for label in ("bull", "base", "bear"):
        s = scenarios.get(label)
        if not s:
            continue

        ws.cell(row=row, column=1, value=label.capitalize()).font = LABEL_FONT
        prob = _safe_float(s.get("probability"))
        fv = _safe_float(s.get("fair_value"))

        prob_col, fv_col, upside_col = 2, 3, 4
        scenario_rows[label] = {"row": row}

        if prob is not None:
            _write_input_cell(ws, row, prob_col, prob / 100 if prob > 1 else prob, fmt=FMT_PCT)
        if fv is not None:
            _write_input_cell(ws, row, fv_col, fv, fmt=FMT_USD_INNER)

        # Upside as formula when possible
        if has_price and fv is not None:
            price_ref = registry.sheet_ref("CurrentPrice")
            _write_formula_cell(ws, row, upside_col,
                                f"=(C{row}-{price_ref})/{price_ref}", fmt=FMT_PCT, font=LINK_FONT)
        else:
            upside = _safe_float(s.get("upside_pct"))
            if upside is None and fv is not None and current_price and current_price > 0:
                upside = ((fv - current_price) / current_price) * 100
            if upside is not None:
                ws.cell(row=row, column=upside_col, value=upside).number_format = FMT_PCT_DISPLAY

        assumptions_text = _format_scenario_assumptions(s.get("assumptions", ""))
        cell = ws.cell(row=row, column=5, value=assumptions_text)
        cell.font = RATIONALE_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        for c in range(1, 6):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Weighted fair value as formula
    if len(scenario_rows) >= 2:
        row += 1
        ws.cell(row=row, column=1, value="Weighted Fair Value").font = LABEL_FONT
        # Build =B4*C4 + B5*C5 + B6*C6
        terms = []
        for label_key in ("bull", "base", "bear"):
            if label_key in scenario_rows:
                r = scenario_rows[label_key]["row"]
                terms.append(f"B{r}*C{r}")
        if terms:
            formula = "=" + "+".join(terms)
            _write_formula_cell(ws, row, 3, formula, fmt=FMT_USD,
                                font=Font(name="Calibri", bold=True, size=12, color=HEADER_COLOR))
        row += 1
    else:
        wfv = _safe_float(scenarios.get("weighted_fair_value"))
        if wfv is not None:
            row += 1
            ws.cell(row=row, column=1, value="Weighted Fair Value").font = LABEL_FONT
            ws.cell(row=row, column=3, value=wfv).number_format = FMT_USD
            ws.cell(row=row, column=3).font = Font(name="Calibri", bold=True, size=12, color=HEADER_COLOR)

    _auto_width(ws)
    ws.column_dimensions["E"].width = 60


# =============================================================================
# EARNINGS MODEL SHEET
# =============================================================================

def _build_earnings_sheet(wb: Workbook, earnings: Dict) -> None:
    ws = wb.create_sheet("Earnings Model")
    ws.cell(row=1, column=1, value="Earnings Model").font = TITLE_FONT
    ws.cell(row=2, column=1, value="(All figures in USD billions unless otherwise stated)").font = LABEL_GRAY_FONT

    rows_data = earnings.get("rows", [])
    if not rows_data:
        return

    row = 3
    _apply_header_row(ws, row, ["Fiscal Year", "Revenue", "Rev Growth", "EBITDA", "EBITDA Margin", "EPS"])
    row += 1

    estimate_fill = PatternFill(
        start_color=_hex_strip(_PALETTE.pdf.estimate_row_bg),
        end_color=_hex_strip(_PALETTE.pdf.estimate_row_bg), fill_type="solid",
    )

    prev_rev_row = None
    for item in rows_data:
        is_est = item.get("is_estimate", False)
        fill = estimate_fill if is_est else None

        fy = item.get("fiscal_year", "")
        label = f"{fy}E" if is_est else str(fy)

        # Col 1: FY label
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = VALUE_FONT
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill

        # Col 2: Revenue (value)
        rev = _safe_float(item.get("revenue"))
        rev_cell = ws.cell(row=row, column=2, value=rev)
        rev_cell.font = VALUE_FONT
        rev_cell.border = THIN_BORDER
        rev_cell.number_format = FMT_BILLIONS
        if fill:
            rev_cell.fill = fill

        # Col 3: Revenue Growth (formula if we have previous row)
        if prev_rev_row and rev is not None:
            formula = f"=IF(B{prev_rev_row}=0,0,(B{row}-B{prev_rev_row})/B{prev_rev_row})"
            _write_formula_cell(ws, row, 3, formula, fmt=FMT_PCT)
        else:
            growth = _safe_float(item.get("revenue_growth"))
            if growth is not None:
                cell3 = ws.cell(row=row, column=3, value=growth)
                cell3.font = VALUE_FONT
                cell3.border = THIN_BORDER
                cell3.number_format = FMT_PCT
        if fill:
            ws.cell(row=row, column=3).fill = fill

        # Col 4: EBITDA (value)
        ebitda = _safe_float(item.get("ebitda"))
        cell4 = ws.cell(row=row, column=4, value=ebitda)
        cell4.font = VALUE_FONT
        cell4.border = THIN_BORDER
        cell4.number_format = FMT_BILLIONS
        if fill:
            cell4.fill = fill

        # Col 5: EBITDA Margin (formula)
        if rev is not None and ebitda is not None:
            formula = f"=IF(B{row}=0,0,D{row}/B{row})"
            _write_formula_cell(ws, row, 5, formula, fmt=FMT_PCT)
        else:
            margin = _safe_float(item.get("ebitda_margin"))
            cell5 = ws.cell(row=row, column=5, value=margin)
            cell5.font = VALUE_FONT
            cell5.border = THIN_BORDER
            cell5.number_format = FMT_PCT
        if fill:
            ws.cell(row=row, column=5).fill = fill

        # Col 6: EPS (value)
        eps = _safe_float(item.get("eps"))
        cell6 = ws.cell(row=row, column=6, value=eps)
        cell6.font = VALUE_FONT
        cell6.border = THIN_BORDER
        cell6.number_format = FMT_USD_INNER
        if fill:
            cell6.fill = fill

        if rev is not None:
            prev_rev_row = row
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Shaded rows = analyst estimates").font = Font(
        name="Calibri", italic=True, size=9, color="888888",
    )

    _auto_width(ws)


# =============================================================================
# COMPARABLE COMPANIES SHEET
# =============================================================================

def _build_comp_sheet(wb: Workbook, comp_data: Dict, registry: Optional[NamedRangeRegistry] = None) -> None:
    ws = wb.create_sheet("Comparable Companies")
    ws.cell(row=1, column=1, value="Comparable Company Analysis").font = TITLE_FONT

    row = 3
    headers = [
        "Company", "Ticker", "Mkt Cap", "P/E", "Fwd P/E",
        "EV/EBITDA", "EV/Rev", "Gross Margin", "Op Margin", "Rev Growth",
    ]
    _apply_header_row(ws, row, headers)
    header_row = row
    row += 1

    median_row: Optional[int] = None

    target = comp_data.get("target", {})
    first_peer_row = row
    if target:
        row = _write_comp_row(ws, row, target, bold=True)
        first_peer_row = row

    peers = comp_data.get("peers", [])
    for peer in peers:
        row = _write_comp_row(ws, row, peer, bold=False)
    last_peer_row = row - 1

    # Medians row with MEDIAN formulas over peer range
    has_peers = len(peers) >= 2 and first_peer_row <= last_peer_row
    if has_peers:
        median_font = Font(name="Calibri", bold=True, italic=True, size=11)
        median_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        ws.cell(row=row, column=1, value="Peer Median").font = median_font
        ws.cell(row=row, column=1).fill = median_fill

        # Columns that get MEDIAN formulas (col 4=P/E through col 10=Rev Growth)
        for col_idx in range(4, 11):
            col = _col_letter(col_idx)
            formula = f"=MEDIAN({col}{first_peer_row}:{col}{last_peer_row})"
            cell = ws.cell(row=row, column=col_idx, value=formula)
            cell.font = median_font
            cell.fill = median_fill
            # Format matches the header
            if col_idx <= 7:
                cell.number_format = FMT_RATIO
            else:
                cell.number_format = FMT_PCT

        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = TOTAL_BORDER
        median_row = row
        row += 2
    else:
        # Fallback to pre-computed medians
        medians = comp_data.get("medians", {})
        if medians:
            ws.cell(row=row, column=1, value="Peer Median").font = Font(
                name="Calibri", bold=True, italic=True, size=11,
            )
            ws.cell(row=row, column=1).fill = PatternFill(
                start_color="F0F0F0", end_color="F0F0F0", fill_type="solid",
            )
            for col, key, fmt in [
                (4, "pe_ratio", FMT_RATIO), (5, "forward_pe", FMT_RATIO),
                (6, "ev_to_ebitda", FMT_RATIO), (7, "ev_to_revenue", FMT_RATIO),
                (8, "gross_margin", FMT_PCT), (9, "operating_margin", FMT_PCT),
                (10, "revenue_growth", FMT_PCT),
            ]:
                val = _safe_float(medians.get(key))
                if val is not None:
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.number_format = fmt
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).border = THIN_BORDER
            median_row = row
            row += 2

    # Implied valuations via formulas referencing median multiples and target data
    implied = comp_data.get("implied_values", {})
    target_financials = comp_data.get("target_financials", {})
    if implied or target_financials:
        ws.cell(row=row, column=1, value="Implied Valuations").font = SUBTITLE_FONT
        row += 1

        # Write target financial metrics as reference cells for formulas
        metric_rows: Dict[str, int] = {}
        fin_items = [
            ("Target EPS", "eps", FMT_USD_INNER),
            ("Target Fwd EPS", "forward_eps", FMT_USD_INNER),
            ("Target Revenue", "revenue", FMT_BILLIONS),
            ("Target EBITDA", "ebitda", FMT_BILLIONS),
        ]
        for label, key, fmt in fin_items:
            val = _safe_float(target_financials.get(key))
            if val is not None:
                ws.cell(row=row, column=1, value=label).font = LABEL_GRAY_FONT
                ws.cell(row=row, column=1).border = THIN_BORDER
                cell = ws.cell(row=row, column=2, value=val)
                cell.number_format = fmt
                cell.font = VALUE_FONT
                cell.border = THIN_BORDER
                metric_rows[key] = row
                row += 1

        # Formula-driven implied share prices: Median Multiple * Target Metric
        # Uses median_row from the comp table above when available
        # P/E multiples give price directly; EV multiples need bridge to equity
        # is_ev: True means Multiple*Metric = EV, need (EV - NetDebt)/Shares
        implied_items = [
            ("P/E Implied", "pe_implied", 4, "eps", False),
            ("Fwd P/E Implied", "forward_pe_implied", 5, "forward_eps", False),
            ("EV/Revenue Implied", "ev_revenue_implied", 7, "revenue", True),
            ("EV/EBITDA Implied", "ev_ebitda_implied", 6, "ebitda", True),
        ]
        has_bridge = registry is not None and registry.has("NetDebt") and registry.has("SharesOutstanding")
        for label, fallback_key, median_col, metric_key, is_ev in implied_items:
            if median_row is not None and metric_key in metric_rows:
                med_cell = f"{_col_letter(median_col)}{median_row}"
                met_cell = f"B{metric_rows[metric_key]}"
                ws.cell(row=row, column=1, value=label).font = LABEL_FONT
                ws.cell(row=row, column=1).border = THIN_BORDER
                if is_ev and has_bridge:
                    net_debt_ref = registry.sheet_ref("NetDebt")
                    shares_ref = registry.sheet_ref("SharesOutstanding")
                    formula = f"=({med_cell}*{met_cell}-{net_debt_ref})/{shares_ref}"
                else:
                    formula = f"={med_cell}*{met_cell}"
                _write_formula_cell(ws, row, 2, formula, fmt=FMT_USD_INNER)
                row += 1
            else:
                val = _safe_float(implied.get(fallback_key))
                if val is not None:
                    _write_label_value(ws, row, label, val, fmt=FMT_USD_INNER)
                    row += 1

    _auto_width(ws)


def _write_comp_row(ws, row: int, company: Dict, bold: bool = False) -> int:
    """Write a single company row in the comp table. Returns next row."""
    font = LABEL_FONT if bold else VALUE_FONT

    ws.cell(row=row, column=1, value=company.get("name", "")).font = font
    ws.cell(row=row, column=2, value=company.get("ticker", "")).font = font

    col_map = [
        (3, "market_cap", FMT_BILLIONS),
        (4, "pe_ratio", FMT_RATIO),
        (5, "forward_pe", FMT_RATIO),
        (6, "ev_to_ebitda", FMT_RATIO),
        (7, "ev_to_revenue", FMT_RATIO),
        (8, "gross_margin", FMT_PCT),
        (9, "operating_margin", FMT_PCT),
        (10, "revenue_growth", FMT_PCT),
    ]
    for col, key, fmt in col_map:
        val = _safe_float(company.get(key))
        if val is not None:
            cell = ws.cell(row=row, column=col, value=val)
            cell.number_format = fmt
            cell.font = font

    for c in range(1, 11):
        ws.cell(row=row, column=c).border = THIN_BORDER

    return row + 1


# =============================================================================
# PRECEDENT TRANSACTIONS SHEET
# =============================================================================

def _build_precedent_sheet(wb: Workbook, precedent: Dict) -> None:
    ws = wb.create_sheet("Precedent Transactions")
    ws.cell(row=1, column=1, value="Precedent Transaction Analysis").font = TITLE_FONT

    deals = precedent.get("deals", [])
    if not deals:
        return

    row = 3
    _apply_header_row(ws, row, [
        "Date", "Acquirer", "Target", "Deal Value (USD)",
        "EV/Revenue", "EV/EBITDA", "Premium %",
    ])
    row += 1

    first_deal_row = row
    for deal in deals:
        ws.cell(row=row, column=1, value=deal.get("date", "")).font = VALUE_FONT
        ws.cell(row=row, column=2, value=deal.get("acquirer", "")).font = VALUE_FONT
        ws.cell(row=row, column=3, value=deal.get("target", "")).font = VALUE_FONT

        for col, key, fmt in [
            (4, "deal_value_usd", FMT_NUMBER),
            (5, "ev_to_revenue", FMT_RATIO),
            (6, "ev_to_ebitda", FMT_RATIO),
            (7, "premium_pct", FMT_PCT),
        ]:
            val = _safe_float(deal.get(key))
            if val is not None:
                ws.cell(row=row, column=col, value=val).number_format = fmt

        for c in range(1, 8):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1
    last_deal_row = row - 1

    # Median metrics as MEDIAN formulas
    row += 1
    ws.cell(row=row, column=1, value="Median Metrics").font = SUBTITLE_FONT
    row += 1

    has_multiple_deals = len(deals) >= 2
    median_pairs = [
        ("Median EV/Revenue", 5, FMT_RATIO),
        ("Median EV/EBITDA", 6, FMT_RATIO),
        ("Median Premium", 7, FMT_PCT),
    ]
    for label, src_col, fmt in median_pairs:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).border = TOTAL_BORDER
        if has_multiple_deals:
            col_l = _col_letter(src_col)
            formula = f"=MEDIAN({col_l}{first_deal_row}:{col_l}{last_deal_row})"
            _write_formula_cell(ws, row, 2, formula, fmt=fmt)
            count_formula = f'=COUNT({col_l}{first_deal_row}:{col_l}{last_deal_row})&" deals"'
            ws.cell(row=row, column=3, value=count_formula).font = LABEL_GRAY_FONT
            ws.cell(row=row, column=3).border = THIN_BORDER
        else:
            # Fallback for the static key names
            fallback_keys = {5: "median_ev_to_revenue", 6: "median_ev_to_ebitda", 7: "median_premium"}
            val = _safe_float(precedent.get(fallback_keys.get(src_col, "")))
            if val is not None:
                _write_label_value(ws, row, label, val, fmt=fmt)
        row += 1

    # Implied values (static, from upstream calculation)
    for label, key, fmt in [
        ("Implied Value (EV/Rev)", "implied_value_ev_revenue", FMT_USD_INNER),
        ("Implied Value (EV/EBITDA)", "implied_value_ev_ebitda", FMT_USD_INNER),
    ]:
        val = _safe_float(precedent.get(key))
        if val is not None:
            _write_label_value(ws, row, label, val, fmt=fmt)
            row += 1

    _auto_width(ws)


# =============================================================================
# CONVICTION SHEET
# =============================================================================

def _build_conviction_sheet(wb: Workbook, conviction: Dict) -> None:
    ws = wb.create_sheet("Conviction")
    ws.cell(row=1, column=1, value="Conviction Score").font = TITLE_FONT

    row = 3
    _write_label_value(ws, row, "Overall Score", _safe_float(conviction.get("overall_score")))
    row += 1
    _write_label_value(ws, row, "Recommendation", conviction.get("recommendation", ""))
    row += 1
    _write_label_value(ws, row, "Confidence", conviction.get("confidence", ""))
    row += 1

    categories = conviction.get("categories", [])
    if categories:
        row += 1
        ws.cell(row=row, column=1, value="Category Breakdown").font = SUBTITLE_FONT
        row += 1
        _apply_header_row(ws, row, ["Category", "Score", "Evidence"])
        row += 1

        for cat in categories:
            ws.cell(row=row, column=1, value=cat.get("name", "")).font = VALUE_FONT
            ws.cell(row=row, column=2, value=_safe_float(cat.get("score"))).font = VALUE_FONT
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")

            evidence = cat.get("evidence", "")
            if isinstance(evidence, list):
                evidence = "; ".join(str(e) for e in evidence)
            cell = ws.cell(row=row, column=3, value=str(evidence)[:500])
            cell.font = VALUE_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            for c in range(1, 4):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

    _auto_width(ws)
    ws.column_dimensions["C"].width = 70


# =============================================================================
# COVER / TABLE OF CONTENTS SHEET
# =============================================================================

def _build_cover_sheet(
    wb: Workbook, ticker: str, company_name: str, rating: str,
    data_as_of: Optional[str] = None,
) -> None:
    """Build a Cover/TOC sheet with hyperlinks to every sheet in the workbook.

    This sheet is built last (after all other sheets exist) but moved to
    position 0 so it appears first when the workbook is opened.
    """
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_properties.tabColor = "808080"

    ws.cell(row=1, column=1, value=company_name).font = Font(
        name="Calibri", bold=True, size=20, color=HEADER_COLOR,
    )
    ws.cell(row=2, column=1, value=f"Ticker: {ticker}").font = SUBTITLE_FONT
    ws.cell(row=3, column=1, value="Equity Research Financial Model").font = VALUE_FONT
    ws.cell(row=4, column=1, value=f"Rating: {rating}").font = LABEL_FONT
    ws.cell(row=5, column=1,
            value=f"Generated: {datetime.now().strftime('%d %B %Y')}").font = VALUE_FONT
    if data_as_of:
        ws.cell(row=6, column=1, value=f"Market Data As Of: {data_as_of}").font = VALUE_FONT
        ws.cell(row=7, column=1, value="George Research LTD").font = LABEL_GRAY_FONT
    else:
        ws.cell(row=6, column=1, value="George Research LTD").font = LABEL_GRAY_FONT

    row = 8 if not data_as_of else 9
    ws.cell(row=row, column=1, value="Table of Contents").font = SUBTITLE_FONT
    row += 1

    _apply_header_row(ws, row, ["Sheet", "Description"])
    row += 1

    sheet_descriptions = {
        "Cover": "This page",
        "Instructions": "Model conventions, color coding, sign conventions",
        "Summary": "Executive summary with key valuation outputs",
        "Assumptions": "All editable inputs (blue cells); single source of truth",
        "Financials": "Historical income statement, balance sheet, cash flows",
        "DCF Model": "Discounted cash flow analysis with formula chain",
        "Sensitivity": "WACC vs TGR and Growth vs Margin sensitivity grids",
        "Scenarios": "Bull / Base / Bear scenario analysis",
        "Earnings Model": "Revenue, EBITDA, and EPS projections",
        "Comparable Companies": "Trading comp analysis with peer medians",
        "Precedent Transactions": "M&A precedent transaction multiples",
        "Conviction": "Conviction score breakdown by category",
        "Checks": "Automated error checks and model integrity dashboard",
    }

    link_font = Font(name="Calibri", size=11, color="0563C1", underline="single")

    for sheet in wb.worksheets:
        if sheet.title == "Cover":
            continue
        cell = ws.cell(row=row, column=1, value=sheet.title)
        cell.font = link_font
        cell.hyperlink = f"#{sheet.title}!A1"
        cell.border = THIN_BORDER

        desc = sheet_descriptions.get(sheet.title, "")
        ws.cell(row=row, column=2, value=desc).font = VALUE_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="CONFIDENTIAL").font = Font(
        name="Calibri", bold=True, size=11, color="C00000",
    )
    ws.cell(row=row + 1, column=1,
            value="This document is for informational purposes only and does not "
            "constitute investment advice.").font = LABEL_GRAY_FONT

    _auto_width(ws)


# =============================================================================
# INSTRUCTIONS SHEET
# =============================================================================

def _build_instructions_sheet(wb: Workbook) -> None:
    """Build an Instructions sheet explaining model conventions.

    Placed as the second tab (after Cover), grey tab color.
    """
    ws = wb.create_sheet("Instructions", 1)
    ws.sheet_properties.tabColor = "808080"

    ws.cell(row=1, column=1, value="Model Instructions").font = TITLE_FONT

    row = 3
    ws.cell(row=row, column=1, value="Color Coding Conventions").font = SUBTITLE_FONT
    row += 1

    color_legend = [
        ("Blue font", "Editable input / assumption",
         INPUT_FONT, None),
        ("Black font", "Formula referencing cells on the same sheet",
         FORMULA_FONT, None),
        ("Green font", "Cross-sheet reference (links to another tab)",
         LINK_FONT, None),
        ("Gray font", "Label or descriptive text",
         LABEL_GRAY_FONT, None),
    ]

    for label, description, font, fill in color_legend:
        cell_a = ws.cell(row=row, column=1, value=label)
        cell_a.font = font
        if fill:
            cell_a.fill = fill
        cell_a.border = THIN_BORDER
        ws.cell(row=row, column=2, value=description).font = VALUE_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Sign Conventions").font = SUBTITLE_FONT
    row += 1
    conventions = [
        "All values are positive on the face of the statements.",
        "Labels indicate direction: (-) prefix for deductions (e.g., '(-) CapEx').",
        "(+) prefix for add-backs (e.g., '(+) D&A Add-back').",
        "Negative numbers are displayed in parentheses: (1,234).",
        "Subtraction is handled by formulas, not by negative input values.",
    ]
    for text in conventions:
        ws.cell(row=row, column=1, value=text).font = VALUE_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Number Format Conventions").font = SUBTITLE_FONT
    row += 1
    formats = [
        ("Currency ($)", "Shown at top and bottom of each schedule only"),
        ("Parentheses (1,234)", "Negative numbers, never minus signs"),
        ("Percentages (0.0%)", "One decimal place"),
        ("Multiples (0.00x)", "Two decimal places with 'x' suffix"),
        ("Billions (B)", "Default unit for financial figures"),
    ]
    for label, desc in formats:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=desc).font = VALUE_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="How to Use This Model").font = SUBTITLE_FONT
    row += 1
    instructions = [
        "1. Navigate to the Assumptions tab to modify inputs (blue cells).",
        "2. All downstream sheets (DCF, Scenarios) recalculate automatically.",
        "3. The Sensitivity tab shows pre-computed grids; change assumptions to re-run.",
        "4. Check the Checks tab for model integrity status before relying on outputs.",
        "5. Protected sheets prevent accidental formula overwrites; only blue cells are editable.",
    ]
    for text in instructions:
        ws.cell(row=row, column=1, value=text).font = VALUE_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Data Sources").font = SUBTITLE_FONT
    row += 1
    sources = [
        ("Financial Statements",
         "Historical data sourced from SEC EDGAR filings and market data providers. "
         "Figures are reported as filed; no manual adjustments."),
        ("Valuation Assumptions",
         "Generated by the George Research analysis engine. "
         "All assumptions are editable on the Assumptions tab."),
        ("Market Data",
         "Share price, market capitalization, and peer multiples reflect data "
         "as of the model generation date."),
        ("Comparable Companies",
         "Peer set selected based on sector, market cap, and business model similarity. "
         "Multiples sourced from market data providers."),
        ("Precedent Transactions",
         "Historical M&A transactions sourced from public deal databases and filings."),
    ]
    for label, desc in sources:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=desc).font = VALUE_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        row += 1

    _auto_width(ws)


# =============================================================================
# SUMMARY SHEET
# =============================================================================

def _build_summary_sheet(
    wb: Workbook, ticker: str, company_name: str, rating: str,
    current_price: Optional[float],
    dcf: Optional[Dict], scenarios: Optional[Dict], football: Optional[Dict],
) -> None:
    ws = wb.create_sheet("Summary")

    ws.cell(row=1, column=1, value=f"{company_name} ({ticker})").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Financial Model Export").font = SUBTITLE_FONT
    ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%d %B %Y')}").font = VALUE_FONT

    row = 5
    _write_label_value(ws, row, "Ticker", ticker)
    row += 1
    _write_label_value(ws, row, "Company", company_name)
    row += 1
    _write_label_value(ws, row, "Rating", rating)
    row += 1
    if current_price is not None:
        _write_label_value(ws, row, "Current Price", current_price, fmt=FMT_USD)
        row += 1

    if dcf:
        # Forward reference: =FairValue is defined on the DCF sheet which is
        # built after Summary. This is safe because Excel resolves named ranges
        # at recalc time, not at write time.
        ws.cell(row=row, column=1, value="DCF Fair Value").font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        _write_formula_cell(ws, row, 2, "=FairValue", fmt=FMT_USD, font=LINK_FONT)
        row += 1
        if current_price and current_price > 0:
            ws.cell(row=row, column=1, value="DCF Upside/Downside").font = LABEL_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            _write_formula_cell(ws, row, 2, "=(FairValue-CurrentPrice)/CurrentPrice",
                                fmt=FMT_PCT, font=LINK_FONT)
            row += 1

    if scenarios:
        wfv = _safe_float(scenarios.get("weighted_fair_value"))
        if wfv is not None:
            _write_label_value(ws, row, "Weighted Fair Value", wfv, fmt=FMT_USD)
            row += 1

    if football:
        ranges = football.get("ranges", [])
        if ranges:
            row += 1
            ws.cell(row=row, column=1, value="Valuation Ranges").font = SUBTITLE_FONT
            row += 1
            _apply_header_row(ws, row, ["Method", "Low", "Mid", "High", "Note"])
            row += 1
            for r in ranges:
                ws.cell(row=row, column=1, value=r.get("method", "")).font = VALUE_FONT
                ws.cell(row=row, column=2, value=_safe_float(r.get("low"))).number_format = FMT_USD_INNER
                ws.cell(row=row, column=3, value=_safe_float(r.get("mid"))).number_format = FMT_USD_INNER
                ws.cell(row=row, column=4, value=_safe_float(r.get("high"))).number_format = FMT_USD_INNER
                source = r.get("source", "")
                if source:
                    ws.cell(row=row, column=5, value=source).font = LABEL_GRAY_FONT
                    ws.cell(row=row, column=5).border = THIN_BORDER
                for c in range(1, 5):
                    ws.cell(row=row, column=c).border = THIN_BORDER
                row += 1

    _auto_width(ws)


# =============================================================================
# PRINT SETUP & FREEZE PANES
# =============================================================================

_LANDSCAPE_SHEETS = {"DCF Model", "Financials", "Comparable Companies",
                     "Precedent Transactions", "Sensitivity", "Earnings Model"}

# Per-sheet freeze panes: freeze below the header/year row so labels and
# headers stay visible when scrolling.  Sheets not listed default to B2.
_FREEZE_PANES = {
    "Cover": None,
    "Instructions": None,
    "DCF Model": "C4",       # Freeze label column + base column, below Y1..Yn headers
    "Financials": "B3",      # Below title rows, before first year column
    "Sensitivity": "B4",     # Below title and subtitle
    "Comparable Companies": "B4",
    "Precedent Transactions": "B4",
    "Earnings Model": "B4",
    "Scenarios": "B4",
}

_PRINT_TITLE_ROWS = {
    "DCF Model": "1:3",
    "Financials": "1:2",
    "Comparable Companies": "1:3",
    "Precedent Transactions": "1:3",
    "Earnings Model": "1:3",
    "Scenarios": "1:3",
}


def _apply_print_setup(ws, company_name: str, ticker: str) -> None:
    """Configure print layout: orientation, fit-to-page, headers/footers, freeze."""
    ws.sheet_properties.pageSetUpPr = ws.sheet_properties.pageSetUpPr or None

    if ws.title in _LANDSCAPE_SHEETS:
        ws.page_setup.orientation = "landscape"
    else:
        ws.page_setup.orientation = "portrait"

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = ws.sheet_properties.pageSetUpPr

    ws.oddHeader.center.text = f"{company_name} ({ticker})"
    ws.oddFooter.left.text = "Confidential"
    ws.oddFooter.right.text = "Page &P of &N"

    ws.print_title_rows = _PRINT_TITLE_ROWS.get(ws.title, "1:1")
    freeze = _FREEZE_PANES.get(ws.title, "B2")
    if freeze:
        ws.freeze_panes = freeze


def _apply_sheet_protection(ws) -> None:
    """Protect sheet with empty password. Only cells with UNLOCKED protection are editable."""
    ws.protection.sheet = True
    ws.protection.password = ""
    ws.protection.enable()


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

def generate_excel(
    session_metadata: Dict[str, Any],
    ticker: str,
    company_name: str,
    rating: str,
    current_price: Optional[float] = None,
    protect_sheets: bool = False,
) -> bytes:
    """
    Generate a formula-driven Excel workbook from session valuation data.

    Args:
        session_metadata: The session.metadata dict containing valuation results.
        ticker: Stock ticker symbol.
        company_name: Full company name.
        rating: Investment recommendation (BUY/HOLD/SELL).
        current_price: Current stock price (for summary calculations).
        protect_sheets: If True, protect all sheets (only Assumptions inputs unlocked).

    Returns:
        XLSX file as bytes.
    """
    wb = Workbook()
    wb.remove(wb.active)

    registry = NamedRangeRegistry()

    dcf = session_metadata.get("dcf")
    sensitivity = session_metadata.get("sensitivity")
    growth_margin = session_metadata.get("sensitivity_operating")
    scenarios = session_metadata.get("scenarios")
    earnings = session_metadata.get("earnings_model")
    precedent = session_metadata.get("precedents")
    football = session_metadata.get("football_field")
    conviction = session_metadata.get("conviction")
    financials = session_metadata.get("financial_statements")
    comp_table = session_metadata.get("comps")
    model_audit = session_metadata.get("model_audit")

    # Tab order: Summary > Assumptions > Financials > DCF > Sensitivity >
    # Scenarios > Earnings > Comps > Precedents > Conviction
    # Cover and Instructions are inserted at positions 0-1 after all sheets
    # exist, so the TOC can hyperlink to every tab.
    _build_summary_sheet(wb, ticker, company_name, rating, current_price, dcf, scenarios, football)

    # Assumptions sheet (must come before DCF so named ranges exist)
    if dcf:
        _build_assumptions_sheet(wb, registry, dcf, current_price,
                                 model_audit=model_audit, scenarios=scenarios)

    if financials and financials.get("is_available"):
        _build_financial_statements_sheet(wb, financials)

    if dcf:
        _build_dcf_sheet(wb, registry, dcf)

    if sensitivity or growth_margin:
        _build_sensitivity_sheet(wb, sensitivity, growth_margin, current_price)

    if scenarios:
        _build_scenarios_sheet(wb, registry, scenarios, current_price)

    if earnings:
        _build_earnings_sheet(wb, earnings)

    if comp_table:
        _build_comp_sheet(wb, comp_table, registry=registry)

    if precedent:
        _build_precedent_sheet(wb, precedent)

    if conviction:
        _build_conviction_sheet(wb, conviction)

    # Professional formatting: tab colors, validation, checks sheet, comments,
    # estimate delineation, row grouping
    from backend.core.excel_formatting import apply_professional_formatting
    apply_professional_formatting(wb, model_audit=model_audit, dcf=dcf)

    # Cover (with TOC) and Instructions must be built last so all sheet
    # titles are available for hyperlinks; they insert at positions 0-1.
    _build_instructions_sheet(wb)
    data_as_of = session_metadata.get("data_as_of") or session_metadata.get("market_data_date")
    _build_cover_sheet(wb, ticker, company_name, rating, data_as_of=data_as_of)

    # Apply print setup, freeze panes, and remove gridlines from all sheets
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        _apply_print_setup(ws, company_name, ticker)
        if protect_sheets:
            _apply_sheet_protection(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
