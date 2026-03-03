"""Tests for formula-driven Excel financial model export.

Generates a workbook from fixture metadata and verifies that key cells
contain formulas (not static values), named ranges are registered,
input cells are unlocked, and the FMT_BILLIONS format is correct.
"""

import io
from typing import Dict

import pytest
from openpyxl import load_workbook

from backend.core.excel_generator import FMT_BILLIONS, generate_excel


# =============================================================================
# Fixtures
# =============================================================================

def _full_metadata() -> Dict:
    """Complete metadata dict with all valuation keys populated."""
    return {
        "dcf": {
            "base_revenue": 400e9,
            "shares_outstanding": 15.5e9,
            "fair_value_per_share": 195.50,
            "current_price": 175.00,
            "enterprise_value": 2800e9,
            "equity_value": 2600e9,
            "net_debt": 50e9,
            "terminal_value": 5000e9,
            "pv_terminal_value": 3200e9,
            "projected_revenues": [440e9, 475e9, 504e9, 529e9, 550e9],
            "projected_fcfs": [88e9, 95e9, 101e9, 106e9, 110e9],
            "discounted_fcfs": [80e9, 79e9, 76e9, 73e9, 69e9],
            "wacc_components": {
                "risk_free_rate": 0.043,
                "beta": 1.15,
                "equity_risk_premium": 0.055,
                "size_premium": 0.0,
                "cost_of_equity": 0.1063,
                "kd_pretax": 0.058,
                "tax_rate": 0.21,
                "kd_aftertax": 0.04582,
                "weight_equity": 0.95,
                "weight_debt": 0.05,
                "wacc": 0.103,
            },
            "ev_bridge": {
                "net_debt": 50e9,
                "minority_interest": 2e9,
            },
            "assumptions": {
                "revenue_growth_rates": [0.10, 0.08, 0.06, 0.05, 0.04],
                "wacc": 0.103,
                "terminal_growth_rate": 0.025,
                "tax_rate": 0.21,
                "gross_margins": [0.45, 0.46, 0.46, 0.47, 0.47],
                "opex_as_pct_revenue": [0.20, 0.19, 0.19, 0.18, 0.18],
                "capex_as_pct_revenue": [0.08, 0.07, 0.07, 0.06, 0.06],
                "da_as_pct_revenue": 0.04,
                "sbc_as_pct_revenue": 0.02,
                "nwc_change_pct": 0.03,
            },
        },
        "scenarios": {
            "bull": {"probability": 25, "fair_value": 250.0},
            "base": {"probability": 50, "fair_value": 195.0},
            "bear": {"probability": 25, "fair_value": 130.0},
            "weighted_fair_value": 192.5,
        },
        "earnings_model": {
            "rows": [
                {"fiscal_year": 2023, "revenue": 380e9, "ebitda": 120e9, "eps": 6.50},
                {"fiscal_year": 2024, "revenue": 400e9, "ebitda": 130e9, "eps": 7.10},
                {"fiscal_year": 2025, "revenue": 440e9, "ebitda": 148e9, "eps": 8.00, "is_estimate": True},
            ],
        },
        "comps": {
            "target": {"name": "Apple", "ticker": "AAPL", "market_cap": 2800e9, "pe_ratio": 28.5},
            "peers": [
                {"name": "MSFT", "ticker": "MSFT", "pe_ratio": 32.0, "forward_pe": 28.0, "ev_to_ebitda": 22.0},
                {"name": "GOOGL", "ticker": "GOOGL", "pe_ratio": 24.0, "forward_pe": 22.0, "ev_to_ebitda": 16.0},
                {"name": "META", "ticker": "META", "pe_ratio": 20.0, "forward_pe": 18.0, "ev_to_ebitda": 14.0},
            ],
            "implied_values": {"pe_implied": 190.0},
        },
        "precedents": {
            "deals": [
                {"date": "2023", "acquirer": "A", "target": "B", "ev_to_revenue": 5.0, "ev_to_ebitda": 15.0},
                {"date": "2022", "acquirer": "C", "target": "D", "ev_to_revenue": 4.0, "ev_to_ebitda": 12.0},
            ],
        },
        "conviction": {
            "overall_score": 7.5,
            "recommendation": "BUY",
            "confidence": "High",
            "categories": [{"name": "Growth", "score": 8, "evidence": "Strong revenue trajectory"}],
        },
        "financial_statements": {
            "is_available": True,
            "income_statements": [
                {"fiscal_year": 2024, "revenue": 400e9, "gross_profit": 180e9,
                 "operating_income": 100e9, "net_income": 80e9},
                {"fiscal_year": 2023, "revenue": 380e9, "gross_profit": 170e9,
                 "operating_income": 95e9, "net_income": 75e9},
            ],
            "balance_sheets": [
                {"fiscal_year": 2024, "total_assets": 500e9, "total_liabilities": 280e9,
                 "total_equity": 220e9, "total_debt": 100e9},
            ],
            "cash_flows": [
                {"fiscal_year": 2024, "operating_cf": 120e9, "free_cash_flow": 90e9},
                {"fiscal_year": 2023, "operating_cf": 110e9, "free_cash_flow": 85e9},
            ],
        },
        "sensitivity": {
            "wacc_values": [0.08, 0.09, 0.10, 0.11, 0.12],
            "terminal_growth_values": [0.01, 0.02, 0.025, 0.03, 0.04],
            "fair_value_grid": [
                [250, 230, 220, 210, 200],
                [220, 200, 190, 180, 170],
                [200, 185, 175, 165, 155],
                [180, 168, 160, 152, 144],
                [165, 155, 148, 140, 132],
            ],
            "base_wacc_idx": 2,
            "base_tgr_idx": 2,
        },
    }


def _load_wb(metadata, **kwargs):
    """Generate Excel and load as openpyxl workbook."""
    xlsx = generate_excel(
        metadata, "AAPL", "Apple Inc", "BUY",
        current_price=175.0, **kwargs,
    )
    return load_workbook(io.BytesIO(xlsx))


def _find_formula(ws, label_match: str, col: int = 2, max_row: int = 60):
    """Find first cell in column whose row label contains label_match and value is a formula."""
    for r in range(1, max_row):
        label = ws.cell(row=r, column=1).value
        val = ws.cell(row=r, column=col).value
        if label and label_match in str(label) and isinstance(val, str) and val.startswith("="):
            return r, val
    return None, None


# =============================================================================
# Tests
# =============================================================================

@pytest.fixture(scope="module")
def workbook():
    return _load_wb(_full_metadata())


class TestDCFFormulas:
    def test_revenue_y1_is_formula_with_growth_ref(self, workbook):
        ws = workbook["DCF Model"]
        _, formula = _find_formula(ws, "Revenue", col=3)
        assert formula is not None, "Revenue Y1 should be a formula"
        assert "Assumptions" in formula, "Revenue Y1 should reference Assumptions sheet"

    def test_terminal_value_is_gordon_growth_formula(self, workbook):
        ws = workbook["DCF Model"]
        _, formula = _find_formula(ws, "Terminal Value")
        assert formula is not None, "Terminal value should be a formula"
        assert "Assumptions" in formula, "TV should reference Assumptions for TGR and WACC"

    def test_equity_value_subtracts_minority_interest(self, workbook):
        ws = workbook["DCF Model"]
        _, formula = _find_formula(ws, "Equity Value")
        assert formula is not None, "Equity Value should be a formula"
        # Should reference at least net debt and minority interest
        assert formula.count("-") >= 2, "Should subtract net debt and minority interest"

    def test_fair_value_per_share_formula(self, workbook):
        ws = workbook["DCF Model"]
        _, formula = _find_formula(ws, "Fair Value / Share")
        assert formula is not None, "Fair value/share should be a formula"
        assert "/" in formula, "Should divide equity value by shares"

    def test_discount_factor_references_wacc(self, workbook):
        ws = workbook["DCF Model"]
        _, formula = _find_formula(ws, "Discount Factor", col=3)
        assert formula is not None
        assert "Assumptions" in formula


class TestAssumptionsSheet:
    def test_wacc_is_formula(self, workbook):
        ws = workbook["Assumptions"]
        _, formula = _find_formula(ws, "WACC")
        assert formula is not None, "WACC should be a formula when wacc_components present"
        assert "Ke" in formula and "WeightEquity" in formula

    def test_cost_of_equity_formula(self, workbook):
        ws = workbook["Assumptions"]
        _, formula = _find_formula(ws, "Cost of Equity")
        assert formula is not None, "Ke should be a formula"
        assert "Rf" in formula and "Beta" in formula and "ERP" in formula

    def test_kd_aftertax_formula(self, workbook):
        ws = workbook["Assumptions"]
        _, formula = _find_formula(ws, "Cost of Debt (after-tax)")
        assert formula is not None
        assert "KdPreTax" in formula and "TaxRate" in formula

    def test_input_cells_are_unlocked(self, workbook):
        ws = workbook["Assumptions"]
        unlocked = 0
        for r in range(1, 50):
            cell = ws.cell(row=r, column=2)
            if cell.protection and not cell.protection.locked:
                unlocked += 1
        assert unlocked >= 10, f"Expected at least 10 unlocked input cells, got {unlocked}"

    def test_input_cells_have_blue_font(self, workbook):
        ws = workbook["Assumptions"]
        blue_count = 0
        for r in range(1, 50):
            cell = ws.cell(row=r, column=2)
            if cell.font and cell.font.color:
                rgb = str(getattr(cell.font.color, "rgb", "") or "")
                if "0000CC" in rgb:
                    blue_count += 1
        assert blue_count >= 5, f"Expected at least 5 blue input cells, got {blue_count}"


class TestNamedRanges:
    def test_core_named_ranges_exist(self, workbook):
        names = set(workbook.defined_names.keys()) if hasattr(workbook.defined_names, "keys") else set()
        expected = {"WACC", "TGR", "BaseRevenue", "SharesOutstanding"}
        missing = expected - names
        assert not missing, f"Missing named ranges: {missing}"

    def test_growth_rate_names_exist(self, workbook):
        names = set(workbook.defined_names.keys())
        for i in range(1, 6):
            assert f"GrowthY{i}" in names, f"Missing GrowthY{i}"

    def test_decomposed_margin_names(self, workbook):
        names = set(workbook.defined_names.keys())
        for prefix in ("GrossMarginY", "OpExPctY", "CapExPctY"):
            assert f"{prefix}1" in names, f"Missing {prefix}1"


class TestFinancialsFormulas:
    def test_gross_margin_is_formula(self, workbook):
        ws = workbook["Financials"]
        _, formula = _find_formula(ws, "Gross Margin")
        assert formula is not None, "Gross margin should be a formula"
        assert "IF" in formula, "Should use IF to guard against division by zero"

    def test_operating_margin_is_formula(self, workbook):
        ws = workbook["Financials"]
        _, formula = _find_formula(ws, "Operating Margin")
        assert formula is not None

    def test_net_margin_is_formula(self, workbook):
        ws = workbook["Financials"]
        _, formula = _find_formula(ws, "Net Margin")
        assert formula is not None

    def test_debt_equity_is_formula(self, workbook):
        ws = workbook["Financials"]
        _, formula = _find_formula(ws, "Debt / Equity")
        assert formula is not None

    def test_no_duplicate_year_labels(self, workbook):
        """Year labels should be deduplicated."""
        ws = workbook["Financials"]
        labels = []
        for r in range(1, 60):
            val = ws.cell(row=r, column=2).value
            if val and str(val).startswith("FY"):
                labels.append(val)
        # Each FY label should appear at most 3 times (income, BS, CF headers)
        from collections import Counter
        counts = Counter(labels)
        for label, count in counts.items():
            assert count <= 3, f"{label} appears {count} times (max 3: income, BS, CF)"


class TestScenariosFormulas:
    def test_weighted_fair_value_is_formula(self, workbook):
        ws = workbook["Scenarios"]
        _, formula = _find_formula(ws, "Weighted", col=3)
        assert formula is not None, "Weighted FV should be a probability-weighted formula"
        assert "+" in formula, "Should sum probability-weighted terms"

    def test_upside_references_current_price(self, workbook):
        ws = workbook["Scenarios"]
        # Find first upside formula
        for r in range(1, 20):
            val = ws.cell(row=r, column=4).value
            if isinstance(val, str) and val.startswith("=") and "Assumptions" in val:
                return  # found cross-sheet price reference
        pytest.fail("Upside formula should reference CurrentPrice from Assumptions")


class TestCompFormulas:
    def test_median_row_uses_median_formula(self, workbook):
        ws = workbook["Comparable Companies"]
        _, formula = _find_formula(ws, "Median", col=4)
        assert formula is not None
        assert "MEDIAN" in formula


class TestPrecedentFormulas:
    def test_median_uses_formula(self, workbook):
        ws = workbook["Precedent Transactions"]
        _, formula = _find_formula(ws, "Median EV/Revenue")
        assert formula is not None
        assert "MEDIAN" in formula


class TestEarningsFormulas:
    def test_revenue_growth_is_formula(self, workbook):
        ws = workbook["Earnings Model"]
        for r in range(1, 20):
            val = ws.cell(row=r, column=3).value
            if isinstance(val, str) and val.startswith("="):
                assert "IF" in val, "Growth formula should guard against div-by-zero"
                return
        pytest.fail("No revenue growth formula found")

    def test_ebitda_margin_is_formula(self, workbook):
        ws = workbook["Earnings Model"]
        for r in range(1, 20):
            val = ws.cell(row=r, column=5).value
            if isinstance(val, str) and val.startswith("="):
                return
        pytest.fail("No EBITDA margin formula found")


class TestFormatAndLayout:
    def test_fmt_billions_has_three_divider_commas(self):
        """FMT_BILLIONS should use three commas to divide by 1B with parenthetical negatives."""
        # Format uses positive;negative pattern with brackets for negatives
        assert ',,,"B"' in FMT_BILLIONS, "Must have three divider commas for billions"
        assert ")" in FMT_BILLIONS, "Must use parenthetical negatives"

    def test_all_sheets_have_freeze_panes(self, workbook):
        # Cover and Instructions intentionally have no freeze panes
        skip = {"Cover", "Instructions"}
        for ws in workbook.worksheets:
            if ws.title in skip:
                continue
            assert ws.freeze_panes is not None, f"{ws.title} missing freeze_panes"

    def test_protection_off_by_default(self, workbook):
        for ws in workbook.worksheets:
            assert not ws.protection.sheet, f"{ws.title} should not be protected by default"

    def test_protection_opt_in(self):
        wb = _load_wb(_full_metadata(), protect_sheets=True)
        for ws in wb.worksheets:
            assert ws.protection.sheet, f"{ws.title} should be protected"


class TestLegacyMode:
    """Legacy mode: no gross_margins, no wacc_components."""

    @pytest.fixture
    def legacy_wb(self):
        metadata = _full_metadata()
        metadata["dcf"]["assumptions"] = {
            "revenue_growth_rates": [0.10, 0.08, 0.06, 0.05, 0.04],
            "wacc": 0.10,
            "terminal_growth_rate": 0.025,
            "fcf_margin": 0.20,
        }
        del metadata["dcf"]["wacc_components"]
        return _load_wb(metadata)

    def test_fcf_equals_revenue_times_margin(self, legacy_wb):
        ws = legacy_wb["DCF Model"]
        _, formula = _find_formula(ws, "Free Cash Flow", col=3)
        assert formula is not None, "Legacy FCF should be a formula"
        assert "Assumptions" in formula, "Should reference FCFMargin from Assumptions"

    def test_no_gross_profit_row(self, legacy_wb):
        ws = legacy_wb["DCF Model"]
        for r in range(1, 30):
            label = ws.cell(row=r, column=1).value
            assert label != "Gross Profit", "Legacy mode should not have Gross Profit row"

    def test_wacc_is_input_not_formula(self, legacy_wb):
        ws = legacy_wb["Assumptions"]
        for r in range(1, 30):
            label = ws.cell(row=r, column=1).value
            val = ws.cell(row=r, column=2).value
            if label and str(label).strip() == "WACC":
                assert not (isinstance(val, str) and val.startswith("=")), \
                    "Legacy WACC should be an input cell, not a formula"
                return
        pytest.fail("WACC not found on Assumptions sheet")


# =============================================================================
# Round 2 Additions: Summary formulas, Checks sheet, discount factor, time sort
# =============================================================================

class TestSummaryFormulas:
    def test_fair_value_is_formula(self, workbook):
        """Summary DCF Fair Value should reference the FairValue named range, not a float."""
        ws = workbook["Summary"]
        for r in range(1, 30):
            label = ws.cell(row=r, column=1).value
            val = ws.cell(row=r, column=2).value
            if label and "DCF Fair Value" in str(label):
                assert isinstance(val, str) and "FairValue" in val, \
                    f"Expected formula with FairValue, got {val}"
                return
        pytest.fail("DCF Fair Value not found on Summary sheet")

    def test_upside_is_formula(self, workbook):
        ws = workbook["Summary"]
        for r in range(1, 30):
            label = ws.cell(row=r, column=1).value
            val = ws.cell(row=r, column=2).value
            if label and "Upside" in str(label):
                assert isinstance(val, str) and val.startswith("="), \
                    f"Expected formula, got {val}"
                return


class TestChecksSheet:
    def test_checks_sheet_exists(self, workbook):
        assert "Checks" in workbook.sheetnames

    def test_checks_has_structural_checks(self, workbook):
        ws = workbook["Checks"]
        all_values = []
        for r in range(1, ws.max_row + 1):
            for c in range(1, 4):
                val = ws.cell(row=r, column=c).value
                if val:
                    all_values.append(str(val))
        # Should contain the TGR vs WACC check (now as a live formula check)
        assert any("TGR" in v for v in all_values), "Should have TGR vs WACC structural check"


class TestChecksSheetWithAudit:
    def test_checks_renders_warnings(self):
        metadata = _full_metadata()
        metadata["model_audit"] = {
            "warnings": [
                {"severity": "critical", "section": "dcf", "message": "WACC too low", "metric": "wacc"},
            ],
            "section_commentary": {},
            "precedent_relevance": [],
            "assumption_notes": {},
        }
        wb = _load_wb(metadata)
        ws = wb["Checks"]
        messages = []
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=3).value
            if val:
                messages.append(val)
        assert any("WACC too low" in str(m) for m in messages)


class TestDiscountFactor:
    def test_discount_factor_less_than_one(self, workbook):
        """Discount factor should be 1/(1+WACC)^n, so values < 1."""
        ws = workbook["DCF Model"]
        for r in range(1, 60):
            label = ws.cell(row=r, column=1).value
            if label and "Discount Factor" in str(label):
                # Check the formula contains 1/(1+...)
                formula = ws.cell(row=r, column=3).value
                assert isinstance(formula, str) and formula.startswith("=")
                assert "1/(1+" in formula or "=1/(1+" in formula, \
                    f"Discount factor formula should be 1/(1+WACC)^n, got: {formula}"
                return
        pytest.fail("Discount Factor row not found")


class TestPVofFCFUsesMultiply:
    def test_pv_fcf_multiplies_by_discount_factor(self, workbook):
        """PV of FCF should use FCF * DF (not FCF / DF)."""
        ws = workbook["DCF Model"]
        for r in range(1, 60):
            label = ws.cell(row=r, column=1).value
            if label and "PV of FCF" in str(label):
                formula = ws.cell(row=r, column=3).value
                assert isinstance(formula, str) and "*" in formula, \
                    f"PV of FCF should multiply, got: {formula}"
                return


class TestFinancialsTimeDirection:
    def test_fiscal_years_ascending(self, workbook):
        """Financial statements should show oldest year on the left."""
        ws = workbook["Financials"]
        # Find the income statement header row
        years = []
        for r in range(1, 30):
            label = ws.cell(row=r, column=1).value
            if label and "Income Statement" in str(label):
                # Next row is the year headers
                for c in range(2, 10):
                    val = ws.cell(row=r + 1, column=c).value
                    if val and str(val).startswith("FY"):
                        try:
                            years.append(int(str(val).replace("FY", "")))
                        except ValueError:
                            pass
                break
        if len(years) >= 2:
            assert years == sorted(years), f"Years should be ascending: {years}"


class TestPerYearDASBCNWC:
    def _defined_names(self, workbook):
        return set(workbook.defined_names)

    def test_da_has_per_year_named_ranges(self, workbook):
        """D&A should have per-year named ranges (DAPctY1, DAPctY2, ...) not a single scalar."""
        names = self._defined_names(workbook)
        assert "DAPctY1" in names, "DAPctY1 named range should exist"
        assert "DAPctY2" in names, "DAPctY2 named range should exist"

    def test_sbc_has_per_year_named_ranges(self, workbook):
        names = self._defined_names(workbook)
        assert "SBCPctY1" in names
        assert "SBCPctY2" in names

    def test_nwc_has_per_year_named_ranges(self, workbook):
        names = self._defined_names(workbook)
        assert "NWCPctY1" in names
        assert "NWCPctY2" in names


class TestDataValidation:
    def test_assumptions_has_data_validation(self, workbook):
        """Assumptions sheet should have at least one data validation rule."""
        ws = workbook["Assumptions"]
        assert len(ws.data_validations.dataValidation) > 0, \
            "Assumptions sheet should have data validation"


class TestUnitHeaders:
    def test_dcf_has_units_subtitle(self, workbook):
        ws = workbook["DCF Model"]
        val = ws.cell(row=2, column=1).value
        assert val and "billion" in str(val).lower(), \
            f"DCF row 2 should have units subtitle, got: {val}"

    def test_financials_has_units_subtitle(self, workbook):
        ws = workbook["Financials"]
        val = ws.cell(row=2, column=1).value
        assert val and "billion" in str(val).lower()

    def test_earnings_has_units_subtitle(self, workbook):
        ws = workbook["Earnings Model"]
        val = ws.cell(row=2, column=1).value
        assert val and "billion" in str(val).lower()
